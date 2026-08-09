#!/usr/bin/env python3
"""
DFIR endpoint-triage parser.

Ingests a Windows host-triage collection (the folder layout produced by the
InfoSec DFIR collector) and produces a risk-scored summary of behavioral and
exfiltration indicators — the analysis that was previously done by hand.

Handles the collector's quirks: UTF-16LE text/CSV artifacts, Chrome/Edge
history SQLite DBs, and Get-ChildItem-style file listings.

Expected package layout (files are matched flexibly by name substring):
    <ROOT>/
      system_info.txt
      system_level_installed_apps.csv        user_level_installed_apps.csv
      running_processes.txt  scheduled_task.txt
      tcp_connections.txt    udp_connections.txt
      wifi_profiles.txt      firewall_settings.csv
      windows_logs/*.evtx
      User_level_files/<User>_files/
          <User> powershell_logs.txt
          <User>_All_files.txt  <User>_Downloads_files.txt  <User>_Documents_files.txt
          <User>_Chrome_Default_History   (SQLite)   <User>_Edge_Default_History (SQLite)

Usage:
    python3 dfir_triage.py /path/to/COLLECTION
    python3 dfir_triage.py /path/to/COLLECTION --json report.json
    python3 dfir_triage.py /path/to/COLLECTION --xlsx report.xlsx

No credentials or network needed. Pure stdlib except optional openpyxl (--xlsx).
"""

from __future__ import annotations

import argparse
import csv
import io
import json
import os
import re
import sqlite3
import sys
import tempfile
from collections import Counter
from datetime import datetime, timezone


# ---------------------------------------------------------------- severities
HIGH, MED, LOW, INFO = "HIGH", "MED", "LOW", "INFO"


class Findings:
    def __init__(self):
        self.items: list[dict] = []
        # Raw datasets captured during parsing for the full-detail XLSX sheets.
        self.browser: list[dict] = []   # {browser, profile, ts, title, url, visits}
        self.commands: list[dict] = []  # {source, ts, command}
        self.downloads: list[dict] = [] # {browser, profile, ts, url, target}
        self.apps: list[dict] = []      # {scope, name, version, publisher, installed}
        self.persistence: list[dict] = []  # {ptype, name, detail, scope, flagged}
        self.network: list[dict] = []   # {source, remote, state, raw}
        self.evtx_events: list[dict] = []   # {log, ts, event_id, severity, description, details}
        self.evtx_summary: list[dict] = []  # {log, provider, event_id, description, severity, count}

    def add(self, category, severity, summary, detail=""):
        self.items.append({
            "category": category, "severity": severity,
            "summary": summary, "detail": detail,
        })

    def add_browser(self, browser, profile, ts, title, url, visits):
        self.browser.append({"browser": browser, "profile": profile, "ts": ts or "",
                             "title": title or "", "url": url or "", "visits": visits})

    def add_command(self, source, command, ts=""):
        cmd = (command or "").strip()
        if cmd:
            self.commands.append({"source": source, "ts": ts or "", "command": cmd})

    def by_sev(self, sev):
        return [f for f in self.items if f["severity"] == sev]


# ---------------------------------------------------------------- indicators
# Installed apps / running processes worth calling out (dual-use, all legit for
# some roles — flagged for the analyst to judge, not auto-incriminating).
REMOTE_ACCESS = ["ultraviewer", "anydesk", "teamviewer", "rustdesk", "screenconnect",
                 "ngrok", "logmein", "gotomypc", "vnc", "remote utilities", "splashtop"]
TRANSFER_TOOLS = ["rclone", "winscp", "filezilla", "cyberduck", "megasync", "mega ",
                  "putty", "mobaxterm", "pscp", "psftp"]
CLOUD_SYNC = ["onedrive", "google drive", "dropbox", "box ", "icloud", "proton drive",
              "mega", "sync.com", "pcloud"]
ELEVATION = ["make me admin", "psexec", "sysinternals", "advanced run", "nircmd"]

# Command-line exfil / staging signatures (PowerShell / cmd history).
EXFIL_CMDS = ["mysqldump", "pg_dump", "pgdump", "mongodump", "sqlcmd", "bcp ",
              "gsutil", "aws s3", "az storage", "rclone", "scp ", "sftp",
              "compress-archive", "robocopy", "7z ", "7za ", "makecab",
              "invoke-webrequest", "invoke-restmethod", "uploadfile", "uploadstring",
              "downloadfile", "downloadstring", "start-bitstransfer", "bitsadmin",
              "certutil -urlcache", "certutil -decode",
              "net use", "new-smbmapping", "curl -t", "curl --upload", "-outfile"]

# Attack/LOLBin & tradecraft signatures (PowerShell / cmd history) — separate
# from exfil so they can be scored distinctly.
LOLBIN_CMDS = ["-encodedcommand", "-enc ", "frombase64string", "iex ", "invoke-expression",
               "mshta", "regsvr32", "rundll32", "wscript", "cscript",
               "add-mppreference -exclusionpath", "set-mppreference -disablerealtime",
               "net user", "net localgroup", "whoami /all", "get-localgroupmember",
               "vssadmin delete", "wevtutil cl", "clear-eventlog", "fsutil usn deletejournal"]

# Lookalike / masquerading process names — deliberate typosquats of legit
# system binaries ONLY. Never list a real binary name here (explorer.exe,
# lsass.exe, svchost.exe are legitimate and must not match).
LOOKALIKE_PROC = ["svch0st", "svchost.exe.", "scvhost", "lssass", "lsasss", "1sass",
                  "csrsss", "rundil32", "rund11", "winlogon.exe.", "taskhostw.exe."]
# Offensive tooling that shouldn't be on a normal endpoint.
OFFENSIVE_PROC = ["mimikatz", "meterpreter", "cobaltstrike", "nmap", "masscan",
                  "psexec", "procdump", "sharphound", "bloodhound"]

# Browser-history keyword buckets.
JOB_SEARCH = ["linkedin.com/jobs", "linkedin.com/messaging", "indeed.com", "glassdoor",
              "ziprecruiter", "dice.com", "monster.com", "greenhouse.io", "lever.co",
              "workday" , "myworkdayjobs", "/careers", "recruit"]
CLOUD_UPLOAD = ["drive.google.com", "dropbox.com", "wetransfer", "mega.nz", "sendspace",
                "filemail", "mediafire", "mail.proton.me", "protonmail", "icloud.com",
                "box.com/s", "smash.", "gofile"]
# Paste / tunnel / anonymizer / webhook infra + raw-code download cradles.
ANON_INFRA = ["pastebin", "paste.ee", "ghostbin", "hastebin", "privatebin", "rentry.co",
              "ngrok", "localtunnel", "serveo", "transfer.sh", "filebin", "anonfiles",
              "temp.sh", "dropmefiles", ".onion", "discord.com/api/webhooks",
              "discordapp.com/api/webhooks", "raw.githubusercontent.com"]
LEGAL_HR = ["lawyer", "attorney", "wrongful", "dismissal", "severance", "employment law",
            "constructive dismissal", "labour board", "labor board", "eeoc"]

# Downloaded / listed files worth surfacing. Each entry is a compiled pattern
# searched against the RAW (case-preserving) name. Most are case-insensitive
# substrings; the PIP acronym is matched case-sensitively and bounded so it hits
# HR "Performance Improvement Plan" docs (PIP.pdf, "John PIP Notice", pip_plan)
# without false-matching lowercase "pip" in picture-in-picture assets, python
# pip, "pipe", or "pipeline".
SENSITIVE_FILE = [
    # resume/CV: match resume/résumé (incl. camelCase like MyResume) but exclude the
    # pause-resume / auto-resume verb forms; "cv" only as a bounded token (not cvss/cvs).
    (re.compile(r"(?<!pause[-_ ])(?<!auto[-_ ])r[eé]sum[eé]s?", re.I), MED, "resume/CV"),
    (re.compile(r"curriculum[ _-]?vitae", re.I), MED, "resume/CV"),
    (re.compile(r"(?:^|[^a-z0-9])cv(?:[^a-z0-9]|$)", re.I), MED, "resume/CV"),
    (re.compile(r"compensation", re.I), MED, "compensation record"),
    (re.compile(r"comp_plan", re.I), MED, "compensation record"),
    (re.compile(r"offer letter", re.I), MED, "offer letter"),
    (re.compile(r"severance", re.I), MED, "severance doc"),
    (re.compile(r"performance improvement", re.I), MED, "PIP"),
    (re.compile(r"pip[ _-]plan", re.I), MED, "PIP"),
    (re.compile(r"(?:^|[^A-Za-z0-9])PIP(?:[^A-Za-z0-9]|$)"), MED, "PIP"),  # acronym, case-sensitive
    (re.compile(r"customer", re.I), HIGH, "possible customer data"),
    (re.compile(r"passport", re.I), LOW, "personal ID"),
]


# Web/image/code asset extensions. "resume"/"PIP" hits on these are almost always
# UI assets (a resume-button icon, picture-in-picture player) rather than HR docs.
ASSET_EXT = re.compile(
    r"\.(?:png|jpe?g|gif|svg|ico|webp|bmp|css|js|mjs|ts|html?|json|woff2?|ttf|eot|map|scss|less)\s*$",
    re.I)


def sensitive_file_match(name: str):
    """First SENSITIVE_FILE pattern matching the (raw) name → (sev, label), else None.

    resume/CV and PIP are document concepts, so a match on a web/image/code asset
    (resume.png, pip.js) is treated as a UI asset and skipped, not a sensitive doc.
    """
    for pat, sev, label in SENSITIVE_FILE:
        if pat.search(name):
            if label in ("resume/CV", "PIP") and ASSET_EXT.search(name):
                continue
            return sev, label
    return None
ARCHIVE_EXT = re.compile(r"\.(zip|7z|rar|tar|gz|bak|dump|dmp|sql|bacpac|mdf|ldf)\b", re.I)
DBDUMP_HINT = re.compile(r"(dump|backup|export|prod|customer|_db|database)", re.I)

# ---- macOS-specific indicators ----
# Shell-history exfil/staging commands (zsh/bash). EXFIL_CMDS above is also applied.
MAC_EXFIL = ["rsync ", "scp ", "sftp ", "curl -t", "curl --upload", "ditto ",
             "tar -c", "tar c", "zip -r", "mysqldump", "pg_dump", "mongodump",
             "aws s3", "gsutil", "rclone"]
# Shell-history LOLBin / tradecraft (macOS).
MAC_TRADECRAFT = ["base64 -d", "base64 --decode", "osascript -e", "nc -l", "ncat ",
                  "socat ", "dscl . -create /users", "dscl . create /users",
                  "launchctl load", "launchctl bootstrap", "defaults write"]
# Highest-signal subset — escalate to HIGH (anti-forensics / security disable).
MAC_TAMPER = ["spctl --master-disable", "csrutil disable", "history -c",
              "rm -rf ~/library/logs", "rm -rf /var/log", "sudo killall -9 falcon"]
# Download-then-execute cradle (curl/wget piped straight into a shell).
MAC_DL_EXEC = re.compile(r"(curl|wget)\b[^\n|]*\|\s*(bash|sh|zsh)\b", re.I)
# Suspicious process/exec paths & patterns on macOS.
MAC_PROC_SUSPECT = ["/tmp/", "/var/folders/", "/private/tmp/", " --headless",
                    "osascript -e"]
# Persistence entries that are legitimate/expected (skip these when flagging).
MAC_PERSIST_NOISE = re.compile(
    r"com\.apple\.|crowdstrike|falcon|jamf|com\.microsoft|keystone|com\.google\.|"
    r"zoom|openvpn|paloalto|globalprotect|1password|com\.docker|nessus|island|"
    r"okta|com\.adobe|logi|citrix|webex|slack|dropbox", re.I)

# ---- Claude Code (.claude) artifacts ----
# Hook / MCP-server commands with download-cradle or covert-channel patterns.
# Hooks run shell commands automatically on agent events — a persistence vector.
CLAUDE_HOT_CMD = re.compile(
    r"curl |wget |base64|/dev/tcp|\bnc -|\bncat\b|\bsocat\b|osascript|"
    r"invoke-webrequest|\biex\b|frombase64string", re.I)
# Approved-permission allowlist entries that imply exfil-capable or destructive
# commands were run and whitelisted by the user.
CLAUDE_RISKY_ALLOW = re.compile(
    r"rm -rf|scp |rsync |sftp|aws s3|gsutil|rclone|mysqldump|pg_dump|mongodump|"
    r"curl -T|--upload|\bncat\b|\bnc -|/dev/tcp|dd if=|history -c|wevtutil|"
    r"certutil|bitsadmin|start-bitstransfer|sudo rm|crontab", re.I)
# Blanket allowlist entries that disable review of whole command classes
# (bare wildcards like "Bash(*)" / "Bash(sudo *)"; specific commands don't count).
CLAUDE_BROAD_ALLOW = re.compile(r"^Bash\(\s*(\*|sudo\s*\*)\s*\)", re.I)


def detect_platform(root: str) -> str:
    """Return 'windows', 'macos', or 'unknown' from marker files."""
    names = set()
    for _dp, _dirs, files in os.walk(root):
        names.update(f.lower() for f in files)
        # only need the top couple of levels for markers
    if any(n in names for n in ("system_info.txt", "scheduled_task.txt")) \
       or os.path.isdir(os.path.join(root, "windows_logs")):
        return "windows"
    if any(n in names for n in ("mac_login_history.txt", "system_launchagents.txt",
                                "system_launchdaemons.txt", "system_information.txt")):
        return "macos"
    return "unknown"


# ---------------------------------------------------------------- helpers
def decode_text(path: str) -> str:
    """Decode a collector text/CSV artifact to clean UTF-8.

    BOM-first (most reliable): PowerShell Out-File defaults to UTF-16 LE (BOM
    \\xff\\xfe). Falls back to UTF-8, then a wide-char heuristic for BOM-less
    UTF-16 LE, then latin-1. Idempotent on already-decoded UTF-8 files.
    """
    raw = open(path, "rb").read()
    if raw[:2] == b"\xff\xfe" or raw[:2] == b"\xfe\xff":
        return raw.decode("utf-16", "replace").replace("\x00", "")
    if raw[:3] == b"\xef\xbb\xbf":
        return raw.decode("utf-8-sig", "replace")
    # No BOM — prefer UTF-8 unless it looks like wide-char mojibake.
    try:
        t = raw.decode("utf-8")
        head = t[:400]
        odd_spaces = sum(1 for i in range(1, len(head), 2) if head[i:i+1] in (" ", "\x00"))
        if odd_spaces < max(len(head) // 4, 1):
            return t
    except UnicodeDecodeError:
        pass
    try:
        return raw.decode("utf-16-le", "replace").replace("\x00", "")
    except Exception:
        return raw.decode("latin-1", "replace").replace("\x00", "")


def read_csv_text(path: str) -> list[dict]:
    # Feed csv pre-split lines: the collector's UTF-16 CSVs carry CRLF/stray-CR
    # terminators that make csv.reader raise "new-line character seen in unquoted
    # field". splitlines() normalizes them; these CSVs have no in-field newlines.
    text = decode_text(path)
    rows = list(csv.DictReader(text.splitlines()))
    return rows


def find(root: str, *substrings, subdir: str | None = None) -> list[str]:
    """Case-insensitive filename-substring search under root (optionally one subdir)."""
    base = os.path.join(root, subdir) if subdir else root
    hits = []
    for dirpath, _dirs, files in os.walk(base):
        for f in files:
            low = f.lower()
            if all(s.lower() in low for s in substrings):
                hits.append(os.path.join(dirpath, f))
    return hits


def one(root, *subs, subdir=None):
    hits = find(root, *subs, subdir=subdir)
    return hits[0] if hits else None


def chrome_time(us) -> str:
    try:
        us = int(us)
        if us <= 0:
            return ""
        return datetime.fromtimestamp(us / 1_000_000 - 11644473600, tz=timezone.utc)\
            .strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        return ""


# ---------------------------------------------------------------- parsers
def parse_system_info(root, F: Findings, platform="windows") -> dict:
    # macOS collectors name it system_information.txt; Windows system_info.txt.
    p = one(root, "system_information") or one(root, "system_info")
    info = {}
    if not p:
        return info
    text = decode_text(p)
    if platform == "macos":
        # system_profiler SPHardwareDataType / SPSoftwareDataType style fields
        fields = ["Computer Name", "Model Name", "Model Identifier", "System Version",
                  "Kernel Version", "Serial Number (system)", "User Name",
                  "Time since boot"]
        for key in fields:
            m = re.search(rf"{re.escape(key)}\s*:\s*(.+)", text)
            if m:
                info[key] = re.sub(r"\s{2,}", " ", m.group(1)).strip()
        info["Host Name"] = info.get("Computer Name", "?")
        info["OS Name"] = info.get("System Version", "?")
        F.add("host", INFO, f"Host {info['Host Name']} — {info['OS Name']}",
              f"model {info.get('Model Identifier','?')}; kernel {info.get('Kernel Version','?')}; "
              f"uptime {info.get('Time since boot','?')}")
        return info
    fields = ["Host Name", "OS Name", "OS Version", "System Boot Time", "Domain",
              "System Manufacturer", "System Model", "Time Zone",
              "Original Install Date"]
    for key in fields:
        m = re.search(rf"{re.escape(key)}\s*:\s*(.+)", text)
        if m:
            info[key] = re.sub(r"\s{2,}", " ", m.group(1)).strip()
    m = re.search(r"Date of Artifact Collection.*?\n\s*\n?\s*(.+)", text, re.S)
    if m:
        info["Collection Time"] = m.group(1).strip().splitlines()[0].strip()
    host = info.get("Host Name", "?")
    F.add("host", INFO, f"Host {host} — {info.get('OS Name','?')}",
          f"Domain {info.get('Domain','?')}; boot {info.get('System Boot Time','?')}; "
          f"model {info.get('System Model','?')}; collected {info.get('Collection Time','?')}")
    return info


def parse_installed_apps(root, F: Findings):
    for scope, subs in (("system", ("system_level_installed_apps",)),
                        ("user", ("user_level_installed_apps",))):
        p = one(root, *subs)
        if not p:
            continue
        try:
            rows = read_csv_text(p)
        except Exception:
            continue
        def field(key):  # strip NULs/control chars that survive some UTF-16 rows
            return re.sub(r"[\x00-\x1f]", "", (r.get(key) or "")).strip()
        for r in rows:
            name = field("DisplayName")
            if not name:
                continue
            low = name.lower()
            when = field("InstallDate")
            # Full inventory for the Installed Applications sheet.
            F.apps.append({"scope": scope, "name": name,
                           "version": field("DisplayVersion"),
                           "publisher": field("Publisher"),
                           "installed": when})
            for bucket, sev, label in (
                (REMOTE_ACCESS, MED, "remote-access tool"),
                (TRANSFER_TOOLS, LOW, "file-transfer/SSH tool"),
                (CLOUD_SYNC, LOW, "cloud-sync client"),
                (ELEVATION, MED, "privilege-elevation tool"),
            ):
                if any(k in low for k in bucket):
                    F.add("installed-app", sev, f"{label}: {name}",
                          f"scope={scope}" + (f", installed {when}" if when else ""))
                    break


def parse_processes(root, F: Findings, platform="windows"):
    p = one(root, "running_processes")
    if not p:
        return
    text = decode_text(p).lower()
    for k in REMOTE_ACCESS + TRANSFER_TOOLS + ["rclone", "mega", "keylog"]:
        if k in text:
            F.add("process", MED, f"remote/transfer tool running: '{k}'",
                  "matched in running_processes at collection time")
    for k in OFFENSIVE_PROC:
        if k in text:
            F.add("process", HIGH, f"offensive/attack tool running: '{k}'",
                  "matched in running_processes at collection time")
    if platform == "macos":
        for k in MAC_PROC_SUSPECT:
            if k in text:
                F.add("process", MED, f"process from suspect path/pattern: '{k.strip()}'",
                      "temp-dir execution, headless browser, or inline AppleScript — verify")
    else:
        for k in LOOKALIKE_PROC:
            if k in text:
                F.add("process", HIGH, f"possible masquerading process: '{k.strip()}'",
                      "lookalike of a legit system binary — verify path/signature")


def parse_scheduled(root, F: Findings):
    p = one(root, "scheduled_task")
    if not p:
        return
    text = decode_text(p)
    tasks = re.findall(r"TaskName:\s*(.+)", text)
    STD = re.compile(r"\\Microsoft\\|\\Windows|GoogleUpdate|OneDrive|Adobe|"
                     r"Lenovo|Intel|CrowdStrike|Island|Okta|Nessus|GlobalProtect", re.I)
    suspicious = []
    for t in tasks:
        t = t.strip()
        flagged = not STD.search(t)
        F.persistence.append({"ptype": "ScheduledTask", "name": t, "detail": "",
                              "scope": "system", "flagged": flagged})
        if flagged:
            suspicious.append(t)
    # Fallback: some listings show bare task names (Get-ChildItem) with no "TaskName:".
    if not tasks:
        for line in text.splitlines():
            s = re.sub(r"\s{2,}", " ", line).strip()
            if s and not s.lower().startswith("date of artifact") and "----" not in s \
               and "Mode" not in s and "Directory:" not in s:
                F.persistence.append({"ptype": "ScheduledTask", "name": s, "detail": "",
                                      "scope": "system", "flagged": not STD.search(s)})
    if suspicious:
        F.add("scheduled-task", MED, f"{len(suspicious)} non-standard scheduled task(s)",
              "; ".join(suspicious[:10]))


CONN_STATE = re.compile(r"\b(ESTABLISHED|LISTEN(?:ING)?|CLOSE_WAIT|CLOSE-WAIT|TIME_WAIT|"
                        r"TIME-WAIT|SYN_SENT|SYN-SENT|SYN_RECV|FIN_WAIT_?[12]|LAST_ACK|"
                        r"CLOSING|CLOSED|BOUND)\b", re.I)


def parse_connections(root, F: Findings):
    # Windows: tcp_connections.txt (+udp) ; macOS: active_network_connections.txt
    paths = find(root, "tcp_connections") + find(root, "udp_connections") \
        + find(root, "active_network_connections")
    ext = []
    for p in dict.fromkeys(paths):  # de-dup, preserve order
        text = decode_text(p)
        src = os.path.basename(p)
        for line in text.splitlines():
            s = re.sub(r"\s{2,}", " ", line).strip()
            if not s or s.lower().startswith("date of artifact"):
                continue
            st = CONN_STATE.search(s)
            # a connection row has a state keyword, an lsof arrow, or proto+address
            if not (st or "->" in s or re.search(r"\b(?:TCP|UDP)\b.*[:.]\d", s, re.I)):
                continue
            rm = re.search(r"->\s*(\S+)", s)  # lsof form: local->remote
            remote = rm.group(1) if rm else ""
            F.network.append({"source": src, "remote": remote,
                              "state": st.group(1).upper() if st else "", "raw": s})
            if st and st.group(1).upper() == "ESTABLISHED" \
               and not re.search(r"127\.0\.0\.1|::1|0\.0\.0\.0", s):
                ext.append(s)
    if ext:
        F.add("network", INFO, f"{len(ext)} established external connection(s) at collection",
              "\n".join(ext[:15]))


def _mac_publisher(app: dict) -> str:
    """Best publisher for an SPApplicationsDataType entry: the code-signing org,
    else how it was obtained (apple / identified developer / mac app store)."""
    sb = app.get("signed_by")
    if isinstance(sb, list) and sb:
        pub = re.sub(r"^Developer ID Application:\s*", "", str(sb[0]))
        pub = re.sub(r"^Software Signing$", "Apple", pub)
        return re.sub(r"\s*\([^)]*\)\s*$", "", pub).strip()   # drop trailing (TEAMID)
    return (app.get("obtained_from") or "").replace("_", " ").strip()


def parse_mac_installed_apps(root, F: Findings):
    # Prefer the full JSON inventory (version/publisher/date); fall back to the
    # `ls -l` listing (name only) for packages collected before it was added.
    # SPApplicationsDataType lists everything incl. hundreds of framework/CoreServices
    # helpers — keep only user-facing app dirs so the inventory matches `/Applications`.
    APP_DIR = re.compile(
        r"^(/Applications|/System/Applications|/Users/[^/]+/Applications)"
        r"(/Utilities|/[^/]+\.localized)?$")
    full = one(root, "installed_apps_full.json")
    used_json = False
    if full:
        try:
            apps = json.loads(decode_text(full)).get("SPApplicationsDataType", [])
        except Exception:
            apps = []
        for a in apps:
            name = (a.get("_name") or "").strip()
            scope = os.path.dirname(a.get("path") or "")
            if not name or not APP_DIR.match(scope):
                continue
            F.apps.append({"scope": scope, "name": name,
                           "version": (a.get("version") or "").strip(),
                           "publisher": _mac_publisher(a),
                           "installed": (a.get("lastModified") or "").strip()})
        used_json = bool(apps)

    # installed_apps.txt is `ls -l /Applications` + `/usr/local/bin`. Use it for the
    # /usr/local/bin CLI tools always; for /Applications only when the JSON is absent.
    p = one(root, "installed_apps.txt") or one(root, "installed_apps")
    if not p:
        return
    raw = decode_text(p)
    for line in raw.splitlines():
        if not re.match(r"[dlbc\-]", line) or line.startswith("total"):
            continue
        m = re.search(r"(?:\d{2}:\d{2}|\s\d{4})\s+(.+?)\s*$", line)
        if not m:
            continue
        name = m.group(1).split(" -> ")[0].strip()  # drop symlink targets
        if not name or name in (".", ".."):
            continue
        scope = "/Applications" if name.lower().endswith(".app") else "/usr/local/bin"
        if used_json and scope == "/Applications":
            continue  # richer JSON row already captured
        F.apps.append({"scope": scope, "name": name, "version": "",
                       "publisher": "", "installed": ""})
    text = raw.lower()
    for bucket, sev, label in ((REMOTE_ACCESS, MED, "remote-access tool"),
                               (TRANSFER_TOOLS, LOW, "file-transfer/SSH tool"),
                               (CLOUD_SYNC, LOW, "cloud-sync client"),
                               (OFFENSIVE_PROC, HIGH, "offensive/security tool")):
        for k in bucket:
            if k in text:
                F.add("installed-app", sev, f"{label}: '{k.strip()}'",
                      "matched in installed_apps.txt")


def parse_homebrew(root, F: Findings):
    """Homebrew packages (homebrew_packages.txt): formulae + casks with versions.

    Brew lives outside /Applications so it is invisible to installed_apps.txt.
    File format: '==> Formulae' / '==> Casks' section headers, then 'name ver...'.
    """
    p = one(root, "homebrew_packages")
    if not p:
        return
    section = ""
    for line in decode_text(p).splitlines():
        s = line.strip()
        if s.lower().startswith("date of artifact"):
            break  # collector date footer is last — stop before its date value line
        if not s:
            continue
        if s.startswith("==>"):
            section = "Homebrew cask" if "cask" in s.lower() else "Homebrew formula"
            continue
        if not section:
            continue
        parts = s.split()
        name = parts[0]
        F.apps.append({"scope": section, "name": name,
                       "version": " ".join(parts[1:]), "publisher": "", "installed": ""})
        low = name.lower()
        for bucket, sev, label in ((REMOTE_ACCESS, MED, "remote-access tool"),
                                   (TRANSFER_TOOLS, LOW, "file-transfer/SSH tool"),
                                   (CLOUD_SYNC, LOW, "cloud-sync client"),
                                   (OFFENSIVE_PROC, HIGH, "offensive/security tool")):
            if any(k.strip() == low for k in bucket):
                F.add("installed-app", sev, f"{label} (Homebrew): {name}", section)
                break


def parse_mac_persistence(root, F: Findings, user_dir=None):
    """Flag non-Apple/non-vendor LaunchAgents, LaunchDaemons, and cron entries.

    System-level artifacts are scanned once (user_dir=None); per-user artifacts
    are scanned when user_dir is given — the two modes are disjoint so nothing
    is double-counted.
    """
    if user_dir:
        targets = [("user", one(user_dir, "launchagents")),
                   ("user", one(user_dir, "cron_jobs"))]
    else:
        targets = [("system", one(root, "system_launchagents")),
                   ("system", one(root, "system_launchdaemons")),
                   ("system", one(root, "system_cron_jobs"))]
    for scope, p in targets:
        if not p:
            continue
        kind = ("cron" if "cron" in os.path.basename(p).lower()
                else "LaunchDaemon" if "daemon" in os.path.basename(p).lower()
                else "LaunchAgent")
        text = decode_text(p)
        flagged = []
        for line in text.splitlines():
            s = line.strip()
            if not s or s.startswith("total") or s.lower().startswith("date of artifact"):
                continue
            # a plist entry or a cron command line
            is_entry = re.search(r"\.plist|@reboot|\* \* \* \*|\d+ \* \* \*|/", s)
            if not is_entry:
                continue
            name = s
            m = re.search(r"(?:\d{2}:\d{2}|\s\d{4})\s+(.+?)\s*$", s)  # ls -l → trailing name
            if m:
                name = m.group(1).split(" -> ")[0].strip()
            is_noise = bool(MAC_PERSIST_NOISE.search(s))
            hot = re.search(r"\.plist|/Users/|/tmp/|/var/folders/|@reboot|\* \* \* \*"
                            r"|/private/tmp/|\.sh\b|\.py\b|osascript|curl|wget|base64", s, re.I)
            is_flagged = bool(hot) and not is_noise
            # Full enumeration for the Persistence sheet (every entry, flagged or not).
            F.persistence.append({"ptype": kind, "name": name[:160],
                                  "detail": re.sub(r"\s{2,}", " ", s)[:200],
                                  "scope": scope, "flagged": is_flagged})
            if is_flagged:
                flagged.append(re.sub(r"\s{2,}", " ", s)[:110])
        if flagged:
            # user-writable script paths / download cradles are higher signal
            hot = [l for l in flagged if re.search(r"/tmp/|/var/folders/|curl|wget|base64|osascript", l, re.I)]
            F.add("persistence", MED if not hot else HIGH,
                  f"{len(flagged)} non-standard {kind} entr(y/ies) ({scope})",
                  "\n".join((hot or flagged)[:10]))


def parse_mac_login(root, F: Findings):
    p = one(root, "mac_login_history")
    if not p:
        return
    lines = [l.strip() for l in decode_text(p).splitlines() if l.strip()]
    if lines:
        F.add("login", INFO, f"login history present ({len(lines)} line(s))",
              "\n".join(lines[:12]))


def parse_downloads_listing(root, F: Findings, user_dir):
    p = one(user_dir, "Downloads_files") if user_dir else None
    if not p:
        return
    text = decode_text(p)
    for line in text.splitlines():
        m = sensitive_file_match(line)
        if m:
            sev, label = m
            F.add("downloaded-file", sev, f"{label} in Downloads",
                  re.sub(r"\s{2,}", " ", line).strip()[:120])
        if ARCHIVE_EXT.search(line):
            F.add("downloaded-file", LOW, "archive/DB file in Downloads",
                  re.sub(r"\s{2,}", " ", line).strip()[:120])


def parse_allfiles(root, F: Findings, user_dir):
    p = one(user_dir, "All_files") if user_dir else None
    if not p:
        return
    text = decode_text(p)
    archives = [l for l in text.splitlines() if ARCHIVE_EXT.search(l) and DBDUMP_HINT.search(l)]
    # surface 2024+ archives distinctly (recent staging is more interesting)
    recent = [l for l in archives if re.search(r"202[4-9]", l)]
    if archives:
        F.add("staging", LOW if not recent else MED,
              f"{len(archives)} archive/DB-dump file(s) in profile ({len(recent)} dated 2024+)",
              "\n".join(re.sub(r'\s{2,}', ' ', l).strip()[:110] for l in (recent or archives)[:12]))


def scan_powershell(root, F: Findings, user_dir):
    hits_paths = [p for p in (find(user_dir, "powershell") if user_dir else [])
                  if "claude_folders" not in p.lower()]
    for p in hits_paths:
        text = decode_text(p)
        low = text.lower()
        # Capture the full history (one command per line) for the detail sheet.
        src = os.path.basename(p)
        for line in text.splitlines():
            F.add_command(src, line)
        found = sorted({k for k in EXFIL_CMDS if k in low})
        # ignore the benign official gcloud installer DownloadFile
        if "downloadfile" in found and "cloudsdk" in low and low.count("downloadfile") == 1:
            found = [f for f in found if f != "downloadfile"]
        lolbins = sorted({k.strip() for k in LOLBIN_CMDS if k in low})
        if found:
            F.add("cmd-history", HIGH, "exfil/staging command(s) in PowerShell history",
                  ", ".join(found))
        if lolbins:
            # log-clearing / Defender-tampering are the most alarming subset
            tamper = any(t in lolbins for t in
                         ("wevtutil cl", "clear-eventlog", "set-mppreference -disablerealtime",
                          "add-mppreference -exclusionpath", "vssadmin delete"))
            F.add("cmd-history", HIGH if tamper else MED,
                  "LOLBin / tradecraft command(s) in PowerShell history",
                  ", ".join(lolbins))
        if not found and not lolbins:
            F.add("cmd-history", INFO, "PowerShell history reviewed — no exfil/staging/LOLBin commands",
                  os.path.basename(p))


def scan_shell_history(root, F: Findings, user_dir):
    """macOS zsh/bash history: exfil/staging, tradecraft, anti-forensics, DL-exec."""
    paths = []
    for name in ("zsh_history", "bash_history"):
        paths += find(user_dir, name) if user_dir else []
    for p in paths:
        text = decode_text(p)
        low = text.lower()
        # Capture the full history for the detail sheet. zsh extended-history lines
        # look like ": <epoch>:<elapsed>;<command>" — pull out the timestamp if present.
        src = os.path.basename(p)
        for line in text.splitlines():
            m = re.match(r"^:\s*(\d+):\d+;(.*)$", line)
            if m:
                F.add_command(src, m.group(2), _iso(int(m.group(1))))
            else:
                F.add_command(src, line)
        exfil = sorted({k.strip() for k in (MAC_EXFIL + EXFIL_CMDS) if k in low})
        craft = sorted({k.strip() for k in MAC_TRADECRAFT if k in low})
        tamper = sorted({k.strip() for k in MAC_TAMPER if k in low})
        dl_exec = bool(MAC_DL_EXEC.search(text))
        if exfil:
            F.add("cmd-history", HIGH, "exfil/staging command(s) in shell history",
                  ", ".join(exfil) + f"  [{os.path.basename(p)}]")
        if dl_exec:
            F.add("cmd-history", HIGH, "download-then-execute cradle in shell history",
                  "curl/wget piped into a shell  [%s]" % os.path.basename(p))
        if tamper:
            F.add("cmd-history", HIGH, "anti-forensics / security-disable in shell history",
                  ", ".join(tamper) + f"  [{os.path.basename(p)}]")
        if craft:
            F.add("cmd-history", MED, "tradecraft command(s) in shell history",
                  ", ".join(craft) + f"  [{os.path.basename(p)}]")
        if not (exfil or craft or tamper or dl_exec):
            F.add("cmd-history", INFO, "shell history reviewed — no exfil/tradecraft commands",
                  os.path.basename(p))


def _iso(unix):
    try:
        if not unix or unix <= 0:
            return ""
        return datetime.fromtimestamp(unix, tz=timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        return ""


def query_browser_history(path: str):
    """Return (urls, downloads) from a Chrome/Edge/Safari history SQLite copy.

    Rows are normalized to (url, title, visit_count, ts_iso). `downloads` is
    Chrome/Edge-only: (tab_url, target_path, ts_iso).
    """
    tmp = tempfile.NamedTemporaryFile(suffix=".db", delete=False)
    tmp.write(open(path, "rb").read()); tmp.close()
    urls, dls = [], []
    try:
        con = sqlite3.connect(f"file:{tmp.name}?mode=ro&immutable=1", uri=True)
        cur = con.cursor()
        # Chrome/Edge (Chromium): urls table, WebKit epoch (µs since 1601-01-01)
        try:
            cur.execute("SELECT url, title, visit_count, last_visit_time FROM urls")
            for u, t, vc, lv in cur.fetchall():
                urls.append((u, t, vc, chrome_time(lv)))
            try:
                cur.execute("SELECT tab_url, target_path, start_time FROM downloads")
                dls = [(a, b, chrome_time(c)) for a, b, c in cur.fetchall()]
            except sqlite3.Error:
                pass
        except sqlite3.Error:
            # Safari: history_items + history_visits, CFAbsoluteTime (since 2001-01-01)
            try:
                cur.execute(
                    "SELECT hi.url, MAX(hv.title), COUNT(*), MAX(hv.visit_time) "
                    "FROM history_items hi JOIN history_visits hv "
                    "ON hi.id = hv.history_item GROUP BY hi.id")
                for u, t, vc, vt in cur.fetchall():
                    ts = _iso(vt + 978307200) if vt else ""
                    urls.append((u, t, vc, ts))
            except sqlite3.Error:
                try:
                    cur.execute("SELECT url, '', visit_count, 0 FROM history_items")
                    urls = [(u, t, vc, "") for u, t, vc, _z in cur.fetchall()]
                except sqlite3.Error:
                    pass
        con.close()
    finally:
        os.unlink(tmp.name)
    return urls, dls


def _history_files(user_dir):
    """History DB files across Chrome/Edge/Safari, Win + mac naming; skip .txt exports
    and anything inside the collected .claude folders (history.jsonl etc.)."""
    out = []
    for f in find(user_dir, "history"):
        low = f.lower()
        if low.endswith(".txt") or "claude_folders" in low:
            continue
        out.append(f)
    return out


def scan_browser(root, F: Findings, user_dir):
    if not user_dir:
        return
    for hp in _history_files(user_dir):
        low_name = hp.lower()
        browser = ("Safari" if "safari" in low_name
                   else "Edge" if "edge" in low_name else "Chrome")
        prof = os.path.basename(hp)
        try:
            urls, dls = query_browser_history(hp)
        except Exception as e:
            F.add("browser", INFO, f"{browser} history not parsed ({prof})", str(e)[:80])
            continue
        # Capture the full history + downloads for the detail sheets (not just flagged).
        for (u, t, vc, ts) in urls:
            F.add_browser(browser, prof, ts, t, u, vc)
        for (turl, tgt, ts) in dls:
            F.downloads.append({"browser": browser, "profile": prof, "ts": ts or "",
                                "url": turl or "", "target": tgt or ""})
        for bucket, sev, label in ((JOB_SEARCH, MED, "job-search / networking"),
                                   (CLOUD_UPLOAD, MED, "personal cloud / file-transfer site"),
                                   (ANON_INFRA, HIGH, "paste / tunnel / anonymizer / webhook infra"),
                                   (LEGAL_HR, MED, "employment-legal research")):
            matches = [(u, t, ts) for (u, t, vc, ts) in urls
                       if any(k in (u or "").lower() for k in bucket)]
            if matches:
                matches.sort(key=lambda m: m[2] or "", reverse=True)
                sample = "\n".join(f"{ts}  {u[:90]}" for (u, t, ts) in matches[:8])
                F.add("browser", sev, f"{browser}: {label} — {len(matches)} visit(s)", sample)
        for (turl, tgt, ts) in dls[:400]:
            low = f"{turl} {tgt}".lower()
            if any(k in low for k in ("proton", "wetransfer", "dropbox", "mega.nz")) \
               or sensitive_file_match(tgt or ""):
                F.add("browser-download", LOW, f"{browser} download of interest",
                      f"{ts}  from {(turl or '')[:60]}  -> {(tgt or '')[:70]}")


# ------------------------------------------------- Claude Code (.claude) review
def _load_json(path):
    try:
        return json.loads(decode_text(path))
    except Exception:
        return None


def _msg_texts(msg) -> list[str]:
    """Plain-text blocks of a transcript message (skips tool_result/tool_use)."""
    c = (msg or {}).get("content")
    if isinstance(c, str):
        return [c]
    out = []
    if isinstance(c, list):
        for b in c:
            if isinstance(b, dict) and b.get("type") == "text":
                out.append(b.get("text", ""))
    return out


def _clip(s, n=110):
    return re.sub(r"\s+", " ", s).strip()[:n]


def _claude_scan_settings(F: Findings, path, scope):
    """settings.json / settings.local.json: hooks, risky allowlists, relaxed modes."""
    data = _load_json(path)
    if not isinstance(data, dict):
        return
    label = f"{scope} {os.path.basename(path)}"

    # Hooks — shell commands executed automatically on agent events.
    cmds = []
    for ev, groups in (data.get("hooks") or {}).items():
        if not isinstance(groups, list):
            continue
        for g in groups:
            for h in (g.get("hooks") or []) if isinstance(g, dict) else []:
                c = (h or {}).get("command", "")
                if c:
                    cmds.append(f"{ev}: {c}")
    if cmds:
        hot = [c for c in cmds if CLAUDE_HOT_CMD.search(c)]
        F.add("claude-config", HIGH if hot else MED,
              f"Claude Code hook(s) configured ({label}) — run automatically on agent events",
              "\n".join(_clip(c) for c in cmds[:8]))

    # MCP servers — arbitrary local command execution surface.
    mcp = data.get("mcpServers") or {}
    if isinstance(mcp, dict) and mcp:
        lines = [f"{k}: {_clip(str((v or {}).get('command', v)))}" for k, v in mcp.items()]
        hot = any(CLAUDE_HOT_CMD.search(l) for l in lines)
        F.add("claude-config", HIGH if hot else LOW,
              f"{len(mcp)} MCP server(s) configured ({label})", "\n".join(lines[:8]))

    # Approved-permission allowlist — these patterns were run AND whitelisted.
    allow = [a for a in ((data.get("permissions") or {}).get("allow") or [])
             if isinstance(a, str)]
    risky = [a for a in allow if CLAUDE_RISKY_ALLOW.search(a)]
    broad = [a for a in allow if CLAUDE_BROAD_ALLOW.search(a)]
    if risky:
        F.add("claude-config", MED,
              f"{len(risky)} exfil-capable/destructive approved command pattern(s) ({label})",
              "\n".join(_clip(a) for a in risky[:10]))
    if broad:
        F.add("claude-config", LOW, f"blanket Bash allowlist entr(y/ies) ({label})",
              "\n".join(_clip(a) for a in broad[:5]))

    # Relaxed permission / trust settings.
    relaxed = []
    perms = data.get("permissions") or {}
    if perms.get("defaultMode") in ("bypassPermissions", "acceptEdits", "auto"):
        relaxed.append(f"permissions.defaultMode={perms['defaultMode']}")
    for k in ("skipAutoPermissionPrompt", "enableAllProjectMcpServers",
              "dangerouslySkipPermissions"):
        if data.get(k):
            relaxed.append(f"{k}=true")
    if relaxed:
        F.add("claude-config", LOW, f"relaxed permission setting(s) ({label})",
              ", ".join(relaxed))
    if data.get("apiKeyHelper"):
        F.add("claude-config", MED, f"custom apiKeyHelper configured ({label})",
              _clip(str(data["apiKeyHelper"])))


def _claude_scan_transcripts(F: Findings, projects_dir):
    """Session transcripts: every user prompt + every Bash command Claude ran."""
    sessions = prompts = 0
    cwds, tmin, tmax = set(), None, None
    buckets = {  # bucket -> (severity, description)
        "exfil": (HIGH, "exfil/staging command(s) run via Claude Code Bash tool"),
        "tamper": (HIGH, "anti-forensics / download-exec command(s) run via Claude Code"),
        "craft": (MED, "tradecraft command(s) run via Claude Code Bash tool"),
        "anon": (MED, "paste/tunnel/anonymizer infra referenced in Claude Code sessions"),
    }
    hits = {k: [] for k in buckets}

    def scan_cmd(cmd):
        low = cmd.lower()
        if any(k in low for k in MAC_EXFIL + EXFIL_CMDS):
            hits["exfil"].append(cmd)
        if any(k in low for k in MAC_TAMPER) or MAC_DL_EXEC.search(cmd):
            hits["tamper"].append(cmd)
        if any(k in low for k in MAC_TRADECRAFT):
            hits["craft"].append(cmd)
        if any(k in low for k in ANON_INFRA):
            hits["anon"].append(cmd)

    for dirpath, _dirs, files in os.walk(projects_dir):
        for fn in files:
            if not fn.endswith(".jsonl"):
                continue
            sessions += 1
            try:
                fh = open(os.path.join(dirpath, fn), errors="replace")
            except OSError:
                continue
            with fh:
                for line in fh:
                    try:
                        e = json.loads(line)
                    except Exception:
                        continue
                    ts = e.get("timestamp") or ""
                    if isinstance(ts, str) and ts:
                        tmin = ts if (tmin is None or ts < tmin) else tmin
                        tmax = ts if (tmax is None or ts > tmax) else tmax
                    if e.get("cwd"):
                        cwds.add(e["cwd"])
                    if e.get("type") == "user":
                        for t in _msg_texts(e.get("message")):
                            low = t.lower()
                            # skip harness-injected blocks — not typed by the user
                            if low.lstrip().startswith(("<task-notification>",
                                                        "<local-command-caveat>",
                                                        "<system-reminder>",
                                                        "<command-name>")):
                                continue
                            prompts += 1
                            for k in ANON_INFRA:
                                i = low.find(k)
                                if i >= 0:
                                    hits["anon"].append(
                                        "[prompt] ..." + t[max(0, i - 40):i + 60] + "...")
                                    break
                    elif e.get("type") == "assistant":
                        c = (e.get("message") or {}).get("content")
                        for b in c if isinstance(c, list) else []:
                            if isinstance(b, dict) and b.get("type") == "tool_use" \
                               and b.get("name") == "Bash":
                                cmd = (b.get("input") or {}).get("command", "")
                                if cmd:
                                    scan_cmd(cmd)
                                    proj = os.path.basename(e.get("cwd") or "") or fn[:12]
                                    F.add_command(f"Claude Code: {proj}", cmd,
                                                  ts if isinstance(ts, str) else "")

    if sessions:
        rng = f"{(tmin or '?')[:10]} .. {(tmax or '?')[:10]}"
        F.add("claude-sessions", INFO,
              f"{sessions} Claude Code session transcript(s), {prompts} user prompt(s), {rng}",
              "projects: " + "; ".join(sorted(cwds)[:12]))
    for k, found in hits.items():
        if not found:
            continue
        sev, desc = buckets[k]
        uniq = list(dict.fromkeys(_clip(c) for c in found))
        F.add("claude-sessions", sev, f"{desc} — {len(found)} occurrence(s)",
              "\n".join(uniq[:8]))


def _claude_scan_content(F: Findings, base, scope):
    """Custom skills/commands/agents + memory + hook scripts + shell snapshots:
    flag download-exec cradles, anti-forensics strings, webhook exfil channels."""
    targets = []
    for sub in ("skills", "commands", "agents", "memory", "hooks", "shell-snapshots"):
        d = os.path.join(base, sub)
        if not os.path.isdir(d):
            continue
        for dirpath, _dirs, files in os.walk(d):
            targets += [os.path.join(dirpath, f) for f in files
                        if f.lower().endswith((".md", ".sh", ".py", ".js", ".json"))]
    for name in ("CLAUDE.md", "MEMORY.md"):
        p = os.path.join(base, name)
        if os.path.isfile(p):
            targets.append(p)
    flagged = []
    for p in targets[:2000]:
        try:
            text = decode_text(p)
        except Exception:
            continue
        low = text.lower()
        why = []
        if MAC_DL_EXEC.search(text):
            why.append("download-then-execute cradle")
        if any(k in low for k in MAC_TAMPER):
            why.append("anti-forensics command")
        if "discord.com/api/webhooks" in low or "discordapp.com/api/webhooks" in low:
            why.append("Discord webhook (exfil channel)")
        if why:
            rel = os.path.relpath(p, base)
            flagged.append(f"{rel}: {', '.join(why)}")
    if flagged:
        F.add("claude-content", MED,
              f"{len(flagged)} suspicious file(s) in {scope} .claude content",
              "\n".join(flagged[:10]))


def parse_claude_desktop(root, F: Findings, user_dir):
    """Review collected Claude Desktop artifacts (<user>_claude_desktop/).

    Credential files (Cookies / buddy-tokens.json / Local State) are excluded by
    the collector, and secret values inside config.json are redacted. Mirrors the
    .claude config checks: MCP servers, installed desktop extensions, relaxed
    permission settings, and a content scan of any local-agent session data.
    """
    if not user_dir:
        return
    cd = None
    for d in sorted(os.listdir(user_dir)):
        if d.lower().endswith("_claude_desktop") and \
           os.path.isdir(os.path.join(user_dir, d)):
            cd = os.path.join(user_dir, d)
            break
    if not cd:
        return
    F.add("claude-desktop", INFO, "Claude Desktop artifacts collected",
          "\n".join(sorted(os.listdir(cd))[:15]))

    # MCP servers — arbitrary local command execution surface.
    mcp_p = os.path.join(cd, "claude_desktop_config.json")
    data = _load_json(mcp_p) if os.path.isfile(mcp_p) else None
    if isinstance(data, dict):
        mcp = data.get("mcpServers") or {}
        if isinstance(mcp, dict) and mcp:
            lines = []
            for k, v in mcp.items():
                cmd = (v or {}).get("command", "")
                cmd += " " + " ".join((v or {}).get("args", []) or [])
                lines.append(f"{k}: {_clip(cmd)}")
            hot = any(CLAUDE_HOT_CMD.search(l) for l in lines)
            F.add("claude-desktop", HIGH if hot else LOW,
                  f"{len(mcp)} Claude Desktop MCP server(s) configured", "\n".join(lines[:8]))

    # Relaxed / permission-bypass settings — the permission-gate bypass lives in
    # claude_desktop_config.json's `preferences`; the extension allowlist flag is
    # in config.json. Scan both.
    relaxed = []
    prefs = (data.get("preferences") if isinstance(data, dict) else None) or {}
    for k, v in prefs.items() if isinstance(prefs, dict) else []:
        if "bypasspermission" in k.lower() and (v is True or
                                                 (isinstance(v, dict) and any(v.values()))):
            relaxed.append(f"{k}=bypass-enabled")
    cfg = _load_json(os.path.join(cd, "config.json"))
    if isinstance(cfg, dict):
        for k, v in cfg.items():
            kl = k.lower()
            if "bypasspermission" in kl and (v is True or
                                             (isinstance(v, dict) and any(v.values()))):
                relaxed.append(f"{k}=bypass-enabled")
            if kl.startswith("dxt:allowlistenabled") and v is False:
                relaxed.append("extension allowlist disabled")
    if relaxed:
        F.add("claude-desktop", MED, "relaxed Claude Desktop permission setting(s)",
              "; ".join(dict.fromkeys(relaxed)))

    # Installed desktop extensions (DXT) — code-execution surface.
    ext = _load_json(os.path.join(cd, "extensions-installations.json"))
    names = []
    if isinstance(ext, dict):
        entries = ext.get("extensions")
        if isinstance(entries, dict):
            names = list(entries.keys())
        elif isinstance(entries, list):
            names = [(e.get("name") or e.get("id") or "?") if isinstance(e, dict) else str(e)
                     for e in entries]
    ext_dir = os.path.join(cd, "Claude_Extensions")
    if not names and os.path.isdir(ext_dir):
        names = [d for d in os.listdir(ext_dir)
                 if os.path.isdir(os.path.join(ext_dir, d))]
    if names:
        F.add("claude-desktop", LOW, f"{len(names)} Claude Desktop extension(s) installed",
              ", ".join(_clip(n, 60) for n in names[:12]))

    # Local-agent / cowork session content + extension payloads — cradle/tamper scan.
    for sub, scope in (("local-agent-mode-sessions", "desktop local-agent sessions"),
                       ("claude-code-sessions", "desktop claude-code sessions"),
                       ("Claude_Extensions", "desktop extension")):
        d = os.path.join(cd, sub)
        if os.path.isdir(d):
            _scan_dir_content(F, d, scope)

    # Logs present — surface for manual timeline review (not keyword-scanned here).
    logs = os.path.join(cd, "logs")
    if os.path.isdir(logs):
        lg = [f for f in os.listdir(logs) if f.lower().endswith(".log")]
        if lg:
            F.add("claude-desktop", INFO,
                  f"{len(lg)} Claude Desktop log file(s) collected",
                  "grep mcp.log / coworkd.log for tool-execution timeline: "
                  + ", ".join(sorted(lg)[:12]))


def _scan_dir_content(F: Findings, d, scope):
    """Keyword-scan text-ish files under a directory for cradles/tamper/webhooks."""
    flagged = []
    count = 0
    for dirpath, _dirs, files in os.walk(d):
        for fn in files:
            if not fn.lower().endswith((".md", ".sh", ".py", ".js", ".json", ".txt", ".log")):
                continue
            count += 1
            if count > 2000:
                break
            p = os.path.join(dirpath, fn)
            try:
                text = decode_text(p)
            except Exception:
                continue
            low = text.lower()
            why = []
            if MAC_DL_EXEC.search(text):
                why.append("download-then-execute cradle")
            if any(k in low for k in MAC_TAMPER):
                why.append("anti-forensics command")
            if "discord.com/api/webhooks" in low or "discordapp.com/api/webhooks" in low:
                why.append("Discord webhook (exfil channel)")
            if why:
                flagged.append(f"{os.path.relpath(p, d)}: {', '.join(why)}")
    if flagged:
        F.add("claude-desktop", MED, f"{len(flagged)} suspicious file(s) in {scope}",
              "\n".join(flagged[:10]))


def parse_claude_folders(root, F: Findings, user_dir):
    """Review collected global + project .claude folders (Claude Code artifacts)."""
    if not user_dir:
        return
    cf = None
    for d in sorted(os.listdir(user_dir)):
        if d.lower().endswith("_claude_folders") and \
           os.path.isdir(os.path.join(user_dir, d)):
            cf = os.path.join(user_dir, d)
            break
    if not cf:
        return
    entries = sorted(e for e in os.listdir(cf)
                     if os.path.isdir(os.path.join(cf, e)))
    if not entries:   # empty stub (older collector zips) — nothing to review
        return
    projects = [e for e in entries if e.startswith("project_")]
    F.add("claude-config", INFO,
          f"Claude Code artifacts collected ({len(projects)} project folder(s)"
          f"{' + global' if 'global_claude' in entries else ''})",
          "\n".join(entries[:15]))

    g = os.path.join(cf, "global_claude")
    if os.path.isdir(g):
        for name in ("settings.json", "settings.local.json"):
            p = os.path.join(g, name)
            if os.path.isfile(p):
                _claude_scan_settings(F, p, "global")
        pj = os.path.join(g, "projects")
        if os.path.isdir(pj):
            _claude_scan_transcripts(F, pj)
        _claude_scan_content(F, g, "global")
        envs = [f for f in os.listdir(g) if f.lower().endswith(".env")]
        if envs:
            F.add("claude-config", LOW,
                  "credential/env file(s) stored in global .claude", ", ".join(envs))

    for prj in projects:
        pdir = os.path.join(cf, prj)
        for name in ("settings.json", "settings.local.json"):
            p = os.path.join(pdir, name)
            if os.path.isfile(p):
                _claude_scan_settings(F, p, prj)
        _claude_scan_content(F, pdir, prj)


# Security/System event IDs worth surfacing, with severity. Mirrors parse_evtx.py.
EVTX_KEY = {
    # anti-forensics / tampering
    1102: (HIGH, "Audit Log Cleared"),
    4719: (HIGH, "System Audit Policy Changed"),
    # account & group changes
    4720: (MED, "User Account Created"),
    4726: (MED, "User Account Deleted"),
    4738: (MED, "User Account Changed"),
    4732: (MED, "Member Added to Local Administrators"),
    4728: (MED, "Member Added to Security-Enabled Global Group"),
    4756: (MED, "Member Added to Universal Group"),
    # persistence / execution
    4697: (MED, "Service Installed"),
    7045: (MED, "New Service Installed"),
    4698: (MED, "Scheduled Task Created"),
    4702: (MED, "Scheduled Task Updated"),
    4104: (MED, "PowerShell Script Block Logged"),
    # credential use / lateral movement
    4648: (MED, "Explicit-Credential Logon (RunAs/lateral)"),
    4625: (LOW, "Failed Logon"),
    4740: (LOW, "Account Locked Out"),
    4672: (LOW, "Special-Privilege Logon (admin rights)"),
    4776: (LOW, "NTLM Credential Validation"),
    5140: (LOW, "Network Share Accessed"),
    5145: (LOW, "Network Share Object Access Checked"),
    # service / process activity
    7040: (LOW, "Service Start Type Changed"),
    4688: (LOW, "Process Created"),
    # context (high volume)
    4624: (INFO, "Successful Logon"),
    4634: (INFO, "Logoff"),
}


def note_evtx(root, F: Findings, do_parse=False):
    logs = find(root, ".evtx")
    if not logs:
        return
    names = ", ".join(f"{os.path.basename(l)} ({os.path.getsize(l)//1024}KB)" for l in logs)
    total_mb = sum(os.path.getsize(l) for l in logs) // (1024 * 1024)
    if not do_parse:
        F.add("windows-logs", INFO,
              f"{len(logs)} Windows event log(s) present ({total_mb}MB) — parsing skipped (--no-evtx)",
              names + " — event logs are parsed by default; --no-evtx was passed to skip them")
        return
    try:
        import Evtx.Evtx as evtx
        from lxml import etree
    except ImportError:
        sys.stderr.write(
            "[warn] .evtx logs present but python-evtx is not installed, so the event logs "
            "were NOT parsed. Re-run via:\n"
            "       uv run --with python-evtx --with lxml --with openpyxl python "
            "dfir_triage.py <PKG> --xlsx <out.xlsx>\n")
        F.add("windows-logs", INFO,
              f"{len(logs)} Windows event log(s) present — NOT parsed (python-evtx missing)",
              names + " — `uv run --with python-evtx --with lxml --with openpyxl` to parse, "
                      "or use EvtxECmd / Chainsaw")
        return
    total_mb = sum(os.path.getsize(l) for l in logs) // (1024 * 1024)
    sys.stderr.write(f"[info] parsing {len(logs)} event log(s) (~{total_mb}MB) — this can take "
                     "a few minutes on a large Security.evtx…\n")

    ns = {"e": "http://schemas.microsoft.com/win/2004/08/events/event"}
    # EventData fields worth surfacing per key event (account, source, service, task,
    # command line, script block, share path…). Order controls display order.
    EVTX_FIELDS = ("TargetUserName", "SubjectUserName", "TargetDomainName", "IpAddress",
                   "WorkstationName", "LogonType", "ServiceName", "ServiceFileName",
                   "TaskName", "MemberName", "NewProcessName", "CommandLine",
                   "ParentProcessName", "ProcessName", "ScriptBlockText",
                   "ShareName", "ShareLocalPath", "RelativeTargetName",
                   "AuthenticationPackageName")
    counts = Counter()
    hits = Counter()          # (event_id) -> count, only for key IDs
    counts_by_log = {}        # log basename -> Counter(eid) for the summary sheet
    prov_by = {}              # (log, eid) -> Counter(provider name) — context for bare IDs
    min_ts = max_ts = None
    for lg in logs:
        logname = os.path.basename(lg)
        by_id = counts_by_log.setdefault(logname, Counter())
        try:
            with evtx.Evtx(lg) as log:
                for rec in log.records():
                    try:
                        r = etree.fromstring(rec.xml().encode("utf-8"))
                    except Exception:
                        continue
                    eid_el = r.find(".//e:EventID", ns)
                    if eid_el is None or not (eid_el.text or "").strip().isdigit():
                        continue
                    eid = int(eid_el.text)
                    counts[eid] += 1
                    by_id[eid] += 1
                    prov_el = r.find(".//e:Provider", ns)
                    prov = ""
                    if prov_el is not None:
                        prov = prov_el.get("Name") or prov_el.get("EventSourceName") or ""
                    prov_by.setdefault((logname, eid), Counter())[prov] += 1
                    ts_el = r.find(".//e:TimeCreated", ns)
                    ts = (ts_el.get("SystemTime", "") if ts_el is not None else "")[:19]
                    if ts:
                        min_ts = ts if (min_ts is None or ts < min_ts) else min_ts
                        max_ts = ts if (max_ts is None or ts > max_ts) else max_ts
                    if eid in EVTX_KEY:
                        hits[eid] += 1
                        sev, desc = EVTX_KEY[eid]
                        fields = {d.get("Name"): (d.text or "").strip()
                                  for d in r.findall(".//e:EventData/e:Data", ns)
                                  if d.get("Name") in EVTX_FIELDS
                                  and (d.text or "").strip() not in ("", "-")}
                        # keep EVTX_FIELDS order; truncate long values (e.g. script blocks)
                        details = "; ".join(f"{k}={fields[k][:160]}"
                                            for k in EVTX_FIELDS if k in fields)[:500]
                        F.evtx_events.append({
                            "log": logname, "ts": ts.replace("T", " "), "event_id": eid,
                            "severity": sev, "description": desc, "details": details})
        except Exception as e:
            F.add("windows-logs", INFO, f"could not parse {os.path.basename(lg)}", str(e)[:80])

    # Per-(log, EventID) counts for the summary sheet, with the dominant provider so a
    # bare EventID has context (key security IDs also annotated with desc/severity).
    for logname, by_id in counts_by_log.items():
        for eid, n in by_id.items():
            sev, desc = EVTX_KEY.get(eid, ("", ""))
            provc = prov_by.get((logname, eid))
            provider = provc.most_common(1)[0][0] if provc else ""
            F.evtx_summary.append({"log": logname, "provider": provider, "event_id": eid,
                                   "description": desc, "severity": sev, "count": n})

    total = sum(counts.values())
    F.add("windows-logs", INFO,
          f"{len(logs)} event log(s) parsed — {total} records, range {min_ts} .. {max_ts}",
          names)
    for eid, n in sorted(hits.items(), key=lambda kv: (EVTX_KEY[kv[0]][0], -kv[1])):
        sev, desc = EVTX_KEY[eid]
        F.add("event-log", sev, f"EventID {eid} — {desc}: {n} occurrence(s)",
              "run parse_evtx.py --event-ids %d <windows_logs/> for details" % eid)
    if 4624 not in counts and 4625 not in counts:
        F.add("event-log", INFO, "no logon events (4624/4625) present",
              "Security audit policy likely not configured — absence is NOT proof of no logons")


# Non-user entries under /Users (macOS) that the collector also copies as
# <name>_files but which are not real user profiles — skip them.
NON_USER_PROFILES = {"public_files", "defaultuser0_files", "shared_files",
                     "library_files", ".localized_files"}


def find_user_dirs(root, only=None):
    base = os.path.join(root, "User_level_files")
    if not os.path.isdir(base):
        return []
    out = []
    for d in sorted(os.listdir(base)):
        full = os.path.join(base, d)
        if not os.path.isdir(full) or d.lower() in NON_USER_PROFILES:
            continue
        if only and only.lower() not in d.lower():
            continue
        out.append(full)
    return out


# ---------------------------------------------------------------- output
def render_console(F: Findings, info, out=sys.stdout):
    order = {HIGH: 0, MED: 1, LOW: 2, INFO: 3}
    items = sorted(F.items, key=lambda f: (order[f["severity"]], f["category"]))
    counts = Counter(f["severity"] for f in F.items)
    print("=" * 72, file=out)
    print(f" DFIR TRIAGE — {info.get('Host Name','(unknown host)')}", file=out)
    print("=" * 72, file=out)
    print(f" HIGH={counts[HIGH]}  MED={counts[MED]}  LOW={counts[LOW]}  INFO={counts[INFO]}",
          file=out)
    print(file=out)
    for f in items:
        print(f"[{f['severity']:4s}] {f['category']}: {f['summary']}", file=out)
        if f["detail"]:
            for line in str(f["detail"]).splitlines():
                print(f"        {line}", file=out)
    print("\n" + "-" * 72, file=out)
    print(" Severity: HIGH=active concern · MED=notable behavioral signal · "
          "LOW=present, judge in context · INFO=context", file=out)


XLSX_MAX_ROWS = 50000  # per detail sheet — guardrail against pathologically large histories


def write_xlsx(F: Findings, info, path, note=None):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
    except ImportError:
        sys.stderr.write("[skip] openpyxl not installed; --xlsx ignored\n")
        return

    def clean(v):
        """Strip control chars openpyxl rejects (common in decoded Windows artifacts)."""
        return ILLEGAL_CHARACTERS_RE.sub("", v) if isinstance(v, str) else v
    fills = {HIGH: "F8CBAD", MED: "FFE699", LOW: "C6E0B4", INFO: "D9E1F2"}
    order = {HIGH: 0, MED: 1, LOW: 2, INFO: 3}
    counts = Counter(f["severity"] for f in F.items)
    HDR_FILL = "2E5496"

    def style_header(ws, r, headers):
        for col, name in enumerate(headers, start=1):
            c = ws.cell(r, col, name)
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor=HDR_FILL)
            c.alignment = Alignment(vertical="top")

    def set_widths(ws, widths):
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

    def findings_table(ws, hdr_row, items):
        """Severity/Category/Finding/Detail table, coloured by severity."""
        style_header(ws, hdr_row, ["Severity", "Category", "Finding", "Detail"])
        r = hdr_row
        for f in sorted(items, key=lambda x: (order[x["severity"]], x["category"])):
            r += 1
            for col, v in enumerate((f["severity"], f["category"], f["summary"], f["detail"]),
                                    start=1):
                c = ws.cell(r, col, clean(v))
                c.fill = PatternFill("solid", fgColor=fills[f["severity"]])
                c.alignment = Alignment(wrap_text=True, vertical="top")
        return r

    def data_table(ws, headers, widths, rows):
        """Plain header + rows; wraps every cell. Caps at XLSX_MAX_ROWS with a note."""
        style_header(ws, 1, headers)
        truncated = len(rows) > XLSX_MAX_ROWS
        for i, rowvals in enumerate(rows[:XLSX_MAX_ROWS], start=2):
            for col, v in enumerate(rowvals, start=1):
                c = ws.cell(i, col, clean(v))
                c.alignment = Alignment(wrap_text=True, vertical="top")
        if not rows:
            ws.cell(2, 1, "(none collected in this package)")
        if truncated:
            ws.cell(min(len(rows), XLSX_MAX_ROWS) + 2, 1,
                    f"… {len(rows) - XLSX_MAX_ROWS} more row(s) omitted (capped at {XLSX_MAX_ROWS})")
        ws.freeze_panes = "A2"
        set_widths(ws, widths)

    # ========================= Sheet 1: Summary =========================
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Summary"
    row = [0]

    def banner(text, *, bold=False, size=11, italic=False, fill=None, wrap=False, height=None):
        row[0] += 1
        r = row[0]
        ws.cell(r, 1, clean(text))
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        c = ws.cell(r, 1)
        c.font = Font(bold=bold, size=size, italic=italic)
        c.alignment = Alignment(wrap_text=wrap, vertical="top")
        if fill:
            for col in range(1, 5):
                ws.cell(r, col).fill = PatternFill("solid", fgColor=fill)
        if height:
            ws.row_dimensions[r].height = height

    banner(f"DFIR Triage — {info.get('Host Name','?')}", bold=True, size=14)
    banner(f"HIGH={counts[HIGH]}   MED={counts[MED]}   LOW={counts[LOW]}   "
           f"INFO={counts[INFO]}      ({len(F.items)} findings total)",
           bold=True, size=12, fill="D9E1F2")
    banner("Severity: HIGH=active concern · MED=notable behavioral signal · "
           "LOW=present, judge in context · INFO=context. Dual-use ≠ guilt — "
           "read the flagged command/file, not just the label.",
           italic=True, wrap=True, height=30)
    banner("Sheets: Summary (all findings) · Claude Folder Review · "
           "Full Browser History · Full Command Line History.", italic=True, wrap=True)
    if note:
        banner("NOTE: " + note, bold=True, wrap=True, fill="FFF2CC",
               height=15 * (1 + len(note) // 90))
    row[0] += 2  # spacer + advance to header row
    hdr = row[0]
    findings_table(ws, hdr, F.items)
    ws.freeze_panes = ws.cell(hdr + 1, 1)
    set_widths(ws, [8, 18, 52, 80])

    # ================== Sheet 2: Claude Folder Review ==================
    cs = wb.create_sheet("Claude Folder Review")
    claude_items = [f for f in F.items if f["category"].startswith("claude")]
    cs.cell(1, 1, "Claude Code (.claude) & Claude Desktop artifact review")
    cs.cell(1, 1).font = Font(bold=True, size=13)
    cs.merge_cells("A1:D1")
    if claude_items:
        findings_table(cs, 3, claude_items)
        cs.freeze_panes = cs.cell(4, 1)
    else:
        cs.cell(3, 1, "No Claude Code / Claude Desktop artifacts were collected for this host.")
    set_widths(cs, [8, 20, 52, 90])

    # ================== Sheet 3: Installed Applications ==================
    aps = wb.create_sheet("Installed Applications")
    apps = sorted(F.apps, key=lambda x: (x["scope"], x["name"].lower()))
    data_table(aps, ["Scope", "Name", "Version", "Publisher", "Install Date"],
               [16, 46, 18, 34, 16],
               [(a["scope"], a["name"], a["version"], a["publisher"], a["installed"])
                for a in apps])

    # ================== Sheet 4: Persistence ==================
    ps = wb.create_sheet("Persistence")
    # flagged (non-standard) entries first, then the rest
    persist = sorted(F.persistence, key=lambda x: (not x["flagged"], x["ptype"], x["name"].lower()))
    data_table(ps, ["Type", "Name / Label", "Detail", "Scope", "Flagged"],
               [16, 46, 60, 10, 9],
               [(e["ptype"], e["name"], e["detail"], e["scope"], "yes" if e["flagged"] else "")
                for e in persist])

    # ================== Sheet 5: Network Connections ==================
    ns = wb.create_sheet("Network Connections")
    data_table(ns, ["Source", "State", "Remote", "Connection (raw)"], [26, 14, 26, 90],
               [(c["source"], c["state"], c["remote"], c["raw"]) for c in F.network])

    # ========== Windows Event Log sheets (only when --evtx parsed them) ==========
    if F.evtx_summary or F.evtx_events:
        es = wb.create_sheet("Event Log Summary")
        # annotated security events first (by severity), then everything else by count —
        # so the meaningful rows lead instead of being buried under Application noise.
        summ = sorted(F.evtx_summary,
                      key=lambda x: (order.get(x["severity"], 9), -x["count"]))
        data_table(es, ["Log", "Provider", "Event ID", "Description", "Severity", "Count"],
                   [16, 34, 10, 40, 10, 10],
                   [(e["log"], e["provider"], e["event_id"], e["description"],
                     e["severity"], e["count"]) for e in summ])
        ke = wb.create_sheet("Key Security Events")
        # newest first, then stable-sort by severity → HIGH first, newest within each
        # (so the important events never fall past the row cap)
        kev = sorted(F.evtx_events, key=lambda x: x["ts"], reverse=True)
        kev.sort(key=lambda x: order.get(x["severity"], 9))
        data_table(ke, ["Log", "Timestamp (UTC)", "Event ID", "Severity", "Description", "Details"],
                   [14, 20, 10, 10, 40, 70],
                   [(e["log"], e["ts"], e["event_id"], e["severity"], e["description"], e["details"])
                    for e in kev])

    # ================== Sheet 6: Full Browser History ==================
    bs = wb.create_sheet("Full Browser History")
    brows = sorted(F.browser, key=lambda x: x["ts"], reverse=True)
    data_table(bs, ["Browser", "Profile", "Timestamp (UTC)", "Title", "URL", "Visits"],
               [10, 30, 20, 48, 80, 7],
               [(b["browser"], b["profile"], b["ts"], b["title"], b["url"], b["visits"])
                for b in brows])

    # ================== Sheet 7: Browser Downloads ==================
    ds = wb.create_sheet("Browser Downloads")
    dls = sorted(F.downloads, key=lambda x: x["ts"], reverse=True)
    data_table(ds, ["Browser", "Profile", "Timestamp (UTC)", "Source URL", "Saved To"],
               [10, 30, 20, 60, 70],
               [(d["browser"], d["profile"], d["ts"], d["url"], d["target"]) for d in dls])

    # ============== Sheet 8: Full Command Line History ==============
    ms = wb.create_sheet("Full Command Line History")
    data_table(ms, ["Source", "Timestamp (UTC)", "Command"], [30, 20, 110],
               [(c["source"], c["ts"], c["command"]) for c in F.commands])

    wb.save(path)
    evtx_note = (f" · EvtxKeyEvents {len(F.evtx_events)}" if F.evtx_events else "")
    print(f"XLSX -> {path}  (Summary {len(F.items)} · Claude {len(claude_items)} · "
          f"Apps {len(F.apps)} · Persistence {len(F.persistence)} · "
          f"Network {len(F.network)} · Browser {len(F.browser)} · "
          f"Downloads {len(F.downloads)} · Commands {len(F.commands)}{evtx_note})")


def main():
    ap = argparse.ArgumentParser(
        description="Parse & risk-score a Windows or macOS DFIR triage package")
    ap.add_argument("package", help="Path to the collection root directory")
    ap.add_argument("--json", dest="json_out", help="Write structured findings JSON here")
    ap.add_argument("--xlsx", dest="xlsx_out", help="Write a colour-coded XLSX report here")
    ap.add_argument("--note", help="Analyst context line shown as a summary banner in the "
                                   "XLSX (e.g. reason the findings read the way they do)")
    ap.add_argument("--user", help="Limit per-profile parsing to matching profile folder(s), "
                                   "e.g. --user rmilankov (substring match)")
    ap.add_argument("--evtx", action="store_true",
                    help="(default) Parse Windows .evtx logs for key security events + the "
                         "Event Log Summary / Key Security Events sheets. On by default now; "
                         "kept for back-compat. Needs python-evtx (run via uv — see SKILL.md).")
    ap.add_argument("--no-evtx", action="store_true",
                    help="Skip .evtx parsing (faster; large Security.evtx takes minutes). "
                         "The default is to always parse and review the event logs.")
    ap.add_argument("--platform", choices=("windows", "macos"),
                    help="Force platform (default: auto-detect from marker files)")
    args = ap.parse_args()

    root = os.path.abspath(os.path.expanduser(args.package))
    if not os.path.isdir(root):
        sys.stderr.write(f"ERROR: not a directory: {root}\n"); sys.exit(1)

    platform = args.platform or detect_platform(root)
    F = Findings()
    F.add("scope", INFO, f"platform detected: {platform}", os.path.basename(root))
    if platform == "unknown":
        sys.stderr.write("[warn] could not detect platform from marker files; "
                         "defaulting to windows. Force with --platform.\n")
        platform = "windows"

    info = parse_system_info(root, F, platform=platform)
    parse_connections(root, F)

    if platform == "macos":
        parse_mac_installed_apps(root, F)
        parse_homebrew(root, F)
        parse_processes(root, F, platform="macos")
        parse_mac_persistence(root, F)
        parse_mac_login(root, F)
    else:
        parse_installed_apps(root, F)
        parse_processes(root, F, platform="windows")
        parse_scheduled(root, F)
        note_evtx(root, F, do_parse=not args.no_evtx)

    user_dirs = find_user_dirs(root, only=args.user)
    if not user_dirs:
        msg = (f"no profile folder matching '{args.user}'" if args.user
               else "no per-user profile folders found under User_level_files/")
        F.add("profile", INFO, msg, "")
    for ud in user_dirs:
        who = os.path.basename(ud).replace("_files", "")
        F.add("profile", INFO, f"profile collected: {who}", ud)
        parse_downloads_listing(root, F, ud)
        parse_allfiles(root, F, ud)
        if platform == "macos":
            scan_shell_history(root, F, ud)
            parse_mac_persistence(root, F, user_dir=ud)
        else:
            scan_powershell(root, F, ud)
        scan_browser(root, F, ud)
        parse_claude_folders(root, F, ud)
        parse_claude_desktop(root, F, ud)

    render_console(F, info)

    if args.json_out:
        with open(args.json_out, "w") as f:
            json.dump({"host": info, "findings": F.items}, f, indent=2)
        print(f"\nJSON -> {args.json_out}")
    if args.xlsx_out:
        write_xlsx(F, info, args.xlsx_out, note=args.note)


if __name__ == "__main__":
    main()
