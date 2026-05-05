#!/usr/bin/env python3
"""
Render a markdown investigation report to a formatted Excel workbook.
Requires: pip install openpyxl

Usage:
    python3 md_to_xlsx.py <input.md>                  # writes alongside as .xlsx
    python3 md_to_xlsx.py <input.md> -o <output.xlsx>
"""

import argparse
import os
import re
import sys

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("openpyxl is required: pip install openpyxl")


# ── colour palette ────────────────────────────────────────────────────────────
H1_FILL   = PatternFill("solid", fgColor="1F3864")
H2_FILL   = PatternFill("solid", fgColor="2E5FA3")
H3_FILL   = PatternFill("solid", fgColor="9DC3E6")
TH_FILL   = PatternFill("solid", fgColor="D6E4F0")
ALT_FILL  = PatternFill("solid", fgColor="F0F6FB")

THIN = Side(style="thin", color="B0C4DE")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

H1_FONT   = Font(name="Calibri", bold=True, size=16, color="FFFFFF")
H2_FONT   = Font(name="Calibri", bold=True, size=13, color="FFFFFF")
H3_FONT   = Font(name="Calibri", bold=True, size=11, color="1F3864")
TH_FONT   = Font(name="Calibri", bold=True, size=10, color="1F3864")
BODY_FONT = Font(name="Calibri", size=10)
META_KEY  = Font(name="Calibri", bold=True, size=10)
CODE_FONT = Font(name="Courier New", size=9, color="444444")


def strip_inline(text):
    """Remove markdown bold/code/link markers, keep plain text."""
    text = re.sub(r"\*\*([^*]+)\*\*", r"\1", text)
    text = re.sub(r"`([^`]+)`", r"\1", text)
    text = re.sub(r"\[([^\]]+)\]\([^)]+\)", r"\1", text)
    return text.strip()


def parse_md(md):
    """
    Parse markdown into a list of blocks:
      ("h1", text)
      ("h2", text)
      ("h3", text)
      ("meta", key, value)   — lines like  **Key:** value
      ("para", text)
      ("ul", [item, ...])
      ("ol", [item, ...])
      ("table", [header_cells], [[row_cells], ...])
      ("hr",)
    """
    lines = md.splitlines()
    blocks = []
    i = 0

    def flush_list(kind, items):
        if items:
            blocks.append((kind, list(items)))

    while i < len(lines):
        line = lines[i]

        # headings
        if line.startswith("### "):
            blocks.append(("h3", strip_inline(line[4:])))
            i += 1; continue
        if line.startswith("## "):
            blocks.append(("h2", strip_inline(line[3:])))
            i += 1; continue
        if line.startswith("# "):
            blocks.append(("h1", strip_inline(line[2:])))
            i += 1; continue

        # horizontal rule
        if re.fullmatch(r"-{3,}", line.strip()):
            blocks.append(("hr",))
            i += 1; continue

        # table — collect consecutive pipe lines
        if line.startswith("|"):
            table_lines = []
            while i < len(lines) and lines[i].startswith("|"):
                table_lines.append(lines[i])
                i += 1
            rows = []
            for tl in table_lines:
                cells = [strip_inline(c) for c in tl.strip().strip("|").split("|")]
                if all(re.fullmatch(r":?-+:?", c) for c in cells):
                    continue  # separator row
                rows.append(cells)
            if rows:
                blocks.append(("table", rows[0], rows[1:]))
            continue

        # unordered list — collect consecutive items
        if line.startswith("- "):
            items = []
            while i < len(lines) and lines[i].startswith("- "):
                items.append(strip_inline(lines[i][2:]))
                i += 1
            blocks.append(("ul", items))
            continue

        # ordered list
        m = re.match(r"^\d+\.\s+(.*)", line)
        if m:
            items = []
            while i < len(lines) and re.match(r"^\d+\.\s+", lines[i]):
                items.append(strip_inline(re.match(r"^\d+\.\s+(.*)", lines[i]).group(1)))
                i += 1
            blocks.append(("ol", items))
            continue

        # meta line: **Key:** value
        mm = re.match(r"^\*\*([^*]+):\*\*\s*(.*)", line)
        if mm:
            blocks.append(("meta", mm.group(1).strip(), mm.group(2).strip()))
            i += 1; continue

        # blank / para
        if line.strip() == "":
            i += 1; continue

        blocks.append(("para", strip_inline(line)))
        i += 1

    return blocks


def write_cell(ws, row, col, value, font=None, fill=None, wrap=True, bold=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font or BODY_FONT
    if fill:
        cell.fill = fill
    cell.alignment = Alignment(wrap_text=wrap, vertical="top")
    return cell


def render(blocks, ws, max_col=6):
    row = 1

    def span(r, text, font, fill, height=None):
        nonlocal row
        ws.merge_cells(start_row=r, start_column=1,
                       end_row=r, end_column=max_col)
        cell = ws.cell(row=r, column=1, value=text)
        cell.font = font
        cell.fill = fill
        cell.alignment = Alignment(horizontal="left", vertical="center",
                                   wrap_text=False)
        if height:
            ws.row_dimensions[r].height = height
        row += 1

    for block in blocks:
        kind = block[0]

        if kind == "h1":
            span(row, block[1], H1_FONT, H1_FILL, height=28)

        elif kind == "h2":
            row += 1
            span(row, block[1], H2_FONT, H2_FILL, height=22)

        elif kind == "h3":
            span(row, block[1], H3_FONT, H3_FILL, height=18)

        elif kind == "meta":
            ws.cell(row=row, column=1, value=block[1] + ":").font = META_KEY
            ws.cell(row=row, column=1).alignment = Alignment(vertical="top")
            c = ws.cell(row=row, column=2, value=block[2])
            c.font = BODY_FONT
            c.alignment = Alignment(wrap_text=True, vertical="top")
            ws.merge_cells(start_row=row, start_column=2,
                           end_row=row, end_column=max_col)
            row += 1

        elif kind == "para":
            ws.merge_cells(start_row=row, start_column=1,
                           end_row=row, end_column=max_col)
            c = ws.cell(row=row, column=1, value=block[1])
            c.font = BODY_FONT
            c.alignment = Alignment(wrap_text=True, vertical="top")
            ws.row_dimensions[row].height = 30
            row += 1

        elif kind in ("ul", "ol"):
            for idx, item in enumerate(block[1]):
                prefix = "•  " if kind == "ul" else f"{idx+1}.  "
                ws.merge_cells(start_row=row, start_column=1,
                               end_row=row, end_column=max_col)
                c = ws.cell(row=row, column=1, value=prefix + item)
                c.font = BODY_FONT
                c.alignment = Alignment(wrap_text=True, vertical="top",
                                        indent=1)
                row += 1

        elif kind == "table":
            headers, data_rows = block[1], block[2]
            ncols = max(len(headers), max((len(r) for r in data_rows), default=0))
            # header row
            for col_i, h in enumerate(headers, start=1):
                c = ws.cell(row=row, column=col_i, value=h)
                c.font = TH_FONT
                c.fill = TH_FILL
                c.border = BORDER
                c.alignment = Alignment(wrap_text=True, vertical="top")
            row += 1
            # data rows
            for r_idx, data_row in enumerate(data_rows):
                fill = ALT_FILL if r_idx % 2 else None
                for col_i, val in enumerate(data_row, start=1):
                    c = ws.cell(row=row, column=col_i, value=val)
                    c.font = BODY_FONT
                    if fill:
                        c.fill = fill
                    c.border = BORDER
                    c.alignment = Alignment(wrap_text=True, vertical="top")
                row += 1
            row += 1  # blank after table

        elif kind == "hr":
            row += 1  # just a gap

    return row


def set_column_widths(ws, max_col):
    widths = {}
    for row in ws.iter_rows():
        for cell in row:
            if cell.value and not isinstance(cell, openpyxl.cell.cell.MergedCell):
                col = cell.column
                widths[col] = max(widths.get(col, 10),
                                  min(len(str(cell.value)), 60))
    for col_i in range(1, max_col + 1):
        ws.column_dimensions[get_column_letter(col_i)].width = widths.get(col_i, 20)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("input", help="Path to markdown file")
    ap.add_argument("-o", "--output", help="Output .xlsx path (default: input.xlsx)")
    args = ap.parse_args()

    if not os.path.exists(args.input):
        sys.exit(f"Input not found: {args.input}")

    out_path = args.output or os.path.splitext(args.input)[0] + ".xlsx"
    out_path = os.path.abspath(out_path)

    md = open(args.input).read()
    blocks = parse_md(md)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Investigation Report"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    render(blocks, ws)
    set_column_widths(ws, max_col=6)

    wb.save(out_path)
    print(f"Wrote {out_path}")


if __name__ == "__main__":
    main()
