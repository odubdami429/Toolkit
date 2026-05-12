#!/usr/bin/env python3
"""
Generate a formatted DFIR investigation report (.xlsx) from Claude's JSON findings file.

Usage:
  python3 generate_report.py <findings.json>
  python3 generate_report.py <findings.json> --out <report.xlsx>

Requires: pip install openpyxl
"""
import argparse
import json
import os
import sys
from datetime import datetime, timezone

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("openpyxl is required:  pip install openpyxl")

# ── palette ──────────────────────────────────────────────────────────────────
_f = lambda c: PatternFill("solid", fgColor=c)

FILL_HDR        = _f("1F3864")   # dark navy  – sheet column headers
FILL_SECTION    = _f("2E5FA3")   # mid blue   – section banners on Summary
FILL_SUBHDR     = _f("9DC3E6")   # light blue – sub-headers
FILL_ALT        = _f("EBF3FB")   # very light blue – alternating rows
FILL_FLAGGED    = _f("FFD7D7")   # light red  – flagged findings
FILL_CLEAN      = _f("D6F0DC")   # light green
FILL_NOTABLE    = _f("FFF2CC")   # light yellow
FILL_SUSPICIOUS = _f("FFD7D7")   # light red
FILL_NA         = _f("F2F2F2")   # grey
FILL_HIGH       = _f("FF0000")   # red – high-confidence IOC
FILL_MED        = _f("FF9900")   # orange – medium IOC
FILL_LOW        = _f("FFFF00")   # yellow – low IOC

THIN   = Side(style="thin", color="B0C4DE")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

_font = lambda **kw: Font(name="Calibri", **kw)
FONT_HDR     = _font(bold=True, size=11, color="FFFFFF")
FONT_SECTION = _font(bold=True, size=12, color="FFFFFF")
FONT_META_K  = _font(bold=True, size=10)
FONT_META_V  = _font(size=10)
FONT_BODY    = _font(size=10)
FONT_BODY_B  = _font(bold=True, size=10)
FONT_CODE    = Font(name="Courier New", size=9, color="333333")
FONT_TITLE   = _font(bold=True, size=16, color="1F3864")

WRAP = Alignment(wrap_text=True, vertical="top")
VCTR = Alignment(vertical="center")


# ── helpers ───────────────────────────────────────────────────────────────────

def _cell(ws, row, col, value="", font=None, fill=None, border=None,
          alignment=None, number_format=None):
    c = ws.cell(row=row, column=col, value=value)
    if font:        c.font        = font
    if fill:        c.fill        = fill
    if border:      c.border      = border
    if alignment:   c.alignment   = alignment
    if number_format: c.number_format = number_format
    return c


def _header_row(ws, row, labels, col_start=1):
    for i, label in enumerate(labels, start=col_start):
        _cell(ws, row, i, label, font=FONT_HDR, fill=FILL_HDR,
              border=BORDER, alignment=Alignment(horizontal="center",
                                                  vertical="center", wrap_text=True))


def _set_col_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _enable_filter(ws, row, ncols):
    last = get_column_letter(ncols)
    ws.auto_filter.ref = f"A{row}:{last}{row}"


def _freeze(ws, row=2, col=1):
    ws.freeze_panes = ws.cell(row=row, column=col)


def _verdict_fill(verdict):
    v = (verdict or "").upper()
    if v == "SUSPICIOUS": return FILL_SUSPICIOUS
    if v == "NOTABLE":    return FILL_NOTABLE
    if v == "CLEAN":      return FILL_CLEAN
    return FILL_NA


def _conf_fill(conf):
    c = (conf or "").upper()
    if c == "HIGH":   return FILL_HIGH
    if c == "MEDIUM": return FILL_MED
    if c == "LOW":    return FILL_LOW
    return FILL_NA


def _row_fill(flagged, row_idx):
    if flagged:
        return FILL_FLAGGED
    return FILL_ALT if row_idx % 2 == 0 else None


# ── sheet builders ────────────────────────────────────────────────────────────

def _sheet_summary(wb, data):
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False

    meta = data.get("metadata", {})
    row = 1

    # Title banner
    ws.merge_cells(f"A{row}:G{row}")
    _cell(ws, row, 1, "DFIR Investigation Report",
          font=FONT_TITLE, alignment=Alignment(horizontal="center", vertical="center"))
    ws.row_dimensions[row].height = 32
    row += 1

    ws.merge_cells(f"A{row}:G{row}")
    generated = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")
    _cell(ws, row, 1, f"Generated: {generated}",
          font=_font(size=9, color="888888"),
          alignment=Alignment(horizontal="center", vertical="center"))
    row += 2

    # ── Metadata ──
    ws.merge_cells(f"A{row}:G{row}")
    _cell(ws, row, 1, "CASE METADATA", font=FONT_SECTION, fill=FILL_SECTION,
          alignment=Alignment(horizontal="left", vertical="center", indent=1))
    ws.row_dimensions[row].height = 20
    row += 1

    meta_fields = [
        ("Hostname",          meta.get("hostname", "")),
        ("Platform",          meta.get("platform", "")),
        ("Collection Time",   meta.get("collection_time", "")),
        ("Analyst",           meta.get("analyst", "")),
        ("Incident Context",  meta.get("incident_context", "")),
    ]
    for k, v in meta_fields:
        _cell(ws, row, 1, k, font=FONT_META_K, border=BORDER,
              alignment=Alignment(vertical="center", indent=1))
        ws.merge_cells(f"B{row}:G{row}")
        _cell(ws, row, 2, v, font=FONT_META_V, border=BORDER,
              alignment=Alignment(vertical="center", wrap_text=True))
        row += 1
    row += 1

    # ── Executive Summary ──
    ws.merge_cells(f"A{row}:G{row}")
    _cell(ws, row, 1, "EXECUTIVE SUMMARY", font=FONT_SECTION, fill=FILL_SECTION,
          alignment=Alignment(horizontal="left", vertical="center", indent=1))
    ws.row_dimensions[row].height = 20
    row += 1

    summary_text = data.get("executive_summary", "")
    ws.merge_cells(f"A{row}:G{row}")
    c = ws.cell(row=row, column=1, value=summary_text)
    c.font = FONT_BODY
    c.border = BORDER
    c.alignment = Alignment(wrap_text=True, vertical="top")
    # Estimate height: ~15px per wrapped line at ~120 chars wide
    lines = max(3, len(summary_text) // 120 + summary_text.count('\n') + 1)
    ws.row_dimensions[row].height = max(60, lines * 15)
    row += 2

    # ── Category Verdicts ──
    ws.merge_cells(f"A{row}:G{row}")
    _cell(ws, row, 1, "ANALYSIS VERDICTS", font=FONT_SECTION, fill=FILL_SECTION,
          alignment=Alignment(horizontal="left", vertical="center", indent=1))
    ws.row_dimensions[row].height = 20
    row += 1

    _header_row(ws, row, ["Category", "Verdict"])
    ws.row_dimensions[row].height = 18
    row += 1

    verdicts = data.get("category_verdicts", {})
    category_labels = {
        "network":        "Network Connections",
        "processes":      "Running Processes",
        "persistence":    "Persistence Mechanisms",
        "software":       "Installed Software",
        "logins":         "Login / Auth History",
        "file_activity":  "User File Activity",
        "shell_history":  "Shell / Command History",
        "browser_history":"Browser History",
        "event_logs":     "Windows Event Logs",
        "firewall":       "Firewall Configuration",
    }
    for key, label in category_labels.items():
        verdict = verdicts.get(key, "N/A").upper()
        fill = _verdict_fill(verdict)
        _cell(ws, row, 1, label, font=FONT_BODY, border=BORDER,
              alignment=Alignment(vertical="center", indent=1))
        _cell(ws, row, 2, verdict, font=FONT_BODY_B, fill=fill, border=BORDER,
              alignment=Alignment(horizontal="center", vertical="center"))
        row += 1

    # Column widths
    _set_col_widths(ws, [22, 18, 20, 20, 20, 20, 20])


def _build_data_sheet(wb, title, columns, col_widths, rows_data):
    """
    Generic data sheet builder.
    rows_data: list of dicts, each must have same keys as columns.
    Last two standard keys expected: 'flagged' (bool) and 'notes' (str).
    """
    ws = wb.create_sheet(title=title)
    ws.sheet_view.showGridLines = False

    hdr_labels = [c["label"] for c in columns]
    _header_row(ws, 1, hdr_labels)
    _enable_filter(ws, 1, len(columns))
    _freeze(ws, 2)

    for i, row_dict in enumerate(rows_data, start=2):
        flagged = bool(row_dict.get("flagged", False))
        fill = _row_fill(flagged, i)
        for j, col_def in enumerate(columns, start=1):
            val = row_dict.get(col_def["key"], "")
            is_notes = col_def["key"] == "flagged"
            font = FONT_CODE if col_def.get("code") else FONT_BODY
            if col_def["key"] == "flagged":
                val = "YES" if val else ""
                font = FONT_BODY_B if flagged else FONT_BODY
            c = ws.cell(row=i, column=j, value=val)
            c.font = font
            if fill: c.fill = fill
            c.border = BORDER
            c.alignment = WRAP if col_def.get("wrap") else Alignment(vertical="top")

    _set_col_widths(ws, col_widths)
    return ws


def _sheet_network(wb, data):
    columns = [
        {"key": "process",      "label": "Process"},
        {"key": "user_pid",     "label": "User / PID"},
        {"key": "local",        "label": "Local Address"},
        {"key": "remote_ip",    "label": "Remote IP"},
        {"key": "remote_port",  "label": "Port"},
        {"key": "state",        "label": "State"},
        {"key": "flagged",      "label": "Flagged"},
        {"key": "notes",        "label": "Notes", "wrap": True},
    ]
    _build_data_sheet(wb, "Network", columns, [22, 18, 22, 22, 8, 14, 9, 40],
                      data.get("network", []))


def _sheet_processes(wb, data):
    columns = [
        {"key": "user",    "label": "User"},
        {"key": "pid",     "label": "PID"},
        {"key": "cpu_pct", "label": "CPU %"},
        {"key": "command", "label": "Command / Path", "wrap": True, "code": True},
        {"key": "flagged", "label": "Flagged"},
        {"key": "notes",   "label": "Notes", "wrap": True},
    ]
    _build_data_sheet(wb, "Processes", columns, [18, 8, 8, 60, 9, 40],
                      data.get("processes", []))


def _sheet_persistence(wb, data):
    columns = [
        {"key": "type",             "label": "Type"},
        {"key": "name",             "label": "Name"},
        {"key": "path_or_command",  "label": "Path / Command", "wrap": True, "code": True},
        {"key": "flagged",          "label": "Flagged"},
        {"key": "notes",            "label": "Notes", "wrap": True},
    ]
    _build_data_sheet(wb, "Persistence", columns, [22, 30, 55, 9, 40],
                      data.get("persistence", []))


def _sheet_software(wb, data):
    columns = [
        {"key": "name",         "label": "Name"},
        {"key": "version",      "label": "Version"},
        {"key": "publisher",    "label": "Publisher"},
        {"key": "install_date", "label": "Install Date"},
        {"key": "flagged",      "label": "Flagged"},
        {"key": "notes",        "label": "Notes", "wrap": True},
    ]
    _build_data_sheet(wb, "Software", columns, [35, 15, 30, 14, 9, 40],
                      data.get("software", []))


def _sheet_logins(wb, data):
    columns = [
        {"key": "user",     "label": "User"},
        {"key": "method",   "label": "Method"},
        {"key": "time_in",  "label": "Time In"},
        {"key": "time_out", "label": "Time Out"},
        {"key": "duration", "label": "Duration"},
        {"key": "flagged",  "label": "Flagged"},
        {"key": "notes",    "label": "Notes", "wrap": True},
    ]
    _build_data_sheet(wb, "Logins", columns, [18, 18, 22, 22, 14, 9, 40],
                      data.get("logins", []))


def _sheet_file_activity(wb, data):
    columns = [
        {"key": "user",     "label": "User"},
        {"key": "location", "label": "Location"},
        {"key": "filename", "label": "Filename", "wrap": True, "code": True},
        {"key": "modified", "label": "Modified"},
        {"key": "flagged",  "label": "Flagged"},
        {"key": "notes",    "label": "Notes", "wrap": True},
    ]
    _build_data_sheet(wb, "File Activity", columns, [18, 14, 55, 20, 9, 40],
                      data.get("file_activity", []))


def _sheet_shell_history(wb, data):
    columns = [
        {"key": "user",    "label": "User"},
        {"key": "shell",   "label": "Shell"},
        {"key": "command", "label": "Command", "wrap": True, "code": True},
        {"key": "flagged", "label": "Flagged"},
        {"key": "notes",   "label": "Notes", "wrap": True},
    ]
    _build_data_sheet(wb, "Shell History", columns, [18, 12, 80, 9, 40],
                      data.get("shell_history", []))


def _sheet_browser_history(wb, data):
    columns = [
        {"key": "user",        "label": "User"},
        {"key": "browser",     "label": "Browser"},
        {"key": "url",         "label": "URL", "wrap": True},
        {"key": "visit_count", "label": "Visits"},
        {"key": "last_visit",  "label": "Last Visit"},
        {"key": "flagged",     "label": "Flagged"},
        {"key": "notes",       "label": "Notes", "wrap": True},
    ]
    _build_data_sheet(wb, "Browser History", columns, [18, 10, 70, 8, 22, 9, 35],
                      data.get("browser_history", []))


def _sheet_event_logs(wb, data):
    rows = data.get("event_logs", [])
    if not rows:
        return
    columns = [
        {"key": "timestamp",   "label": "Timestamp"},
        {"key": "event_id",    "label": "Event ID"},
        {"key": "description", "label": "Description"},
        {"key": "details",     "label": "Details", "wrap": True},
        {"key": "flagged",     "label": "Flagged"},
        {"key": "notes",       "label": "Notes", "wrap": True},
    ]
    _build_data_sheet(wb, "Event Logs", columns, [22, 10, 35, 60, 9, 35], rows)


def _sheet_iocs(wb, data):
    iocs = data.get("iocs", [])
    ws = wb.create_sheet(title="IOCs")
    ws.sheet_view.showGridLines = False

    headers = ["Type", "Value", "Context", "Confidence"]
    _header_row(ws, 1, headers)
    _enable_filter(ws, 1, 4)
    _freeze(ws, 2)

    for i, ioc in enumerate(iocs, start=2):
        conf = (ioc.get("confidence") or "").upper()
        conf_fill = _conf_fill(conf)
        fill = _row_fill(True, i) if conf == "HIGH" else _row_fill(False, i)

        _cell(ws, i, 1, ioc.get("type", ""),       font=FONT_BODY,   border=BORDER, fill=fill)
        _cell(ws, i, 2, ioc.get("value", ""),      font=FONT_CODE,   border=BORDER, fill=fill,
              alignment=WRAP)
        _cell(ws, i, 3, ioc.get("context", ""),    font=FONT_BODY,   border=BORDER, fill=fill,
              alignment=WRAP)
        _cell(ws, i, 4, conf,                       font=FONT_BODY_B, border=BORDER, fill=conf_fill,
              alignment=Alignment(horizontal="center", vertical="center"))

    _set_col_widths(ws, [18, 45, 55, 14])


def _sheet_recommendations(wb, data):
    recs = data.get("recommendations", [])
    ws = wb.create_sheet(title="Recommendations")
    ws.sheet_view.showGridLines = False

    headers = ["Priority", "Action", "Details"]
    _header_row(ws, 1, headers)
    _freeze(ws, 2)

    priority_fills = {
        "1": _f("FF0000"), "2": _f("FF9900"),
        "3": _f("FFFF00"), "4": _f("D6F0DC"), "5": _f("F2F2F2"),
    }
    priority_fonts = {
        "1": _font(bold=True, size=10, color="FFFFFF"),
        "2": _font(bold=True, size=10, color="FFFFFF"),
        "3": _font(bold=True, size=10, color="333333"),
        "4": _font(bold=True, size=10, color="333333"),
        "5": _font(bold=True, size=10, color="333333"),
    }

    for i, rec in enumerate(recs, start=2):
        p = str(rec.get("priority", ""))
        pfill = priority_fills.get(p, FILL_NA)
        pfont = priority_fonts.get(p, FONT_BODY_B)
        row_fill = _row_fill(False, i)

        _cell(ws, i, 1, p, font=pfont, fill=pfill, border=BORDER,
              alignment=Alignment(horizontal="center", vertical="center"))
        _cell(ws, i, 2, rec.get("action", ""), font=FONT_BODY_B,
              fill=row_fill, border=BORDER, alignment=WRAP)
        _cell(ws, i, 3, rec.get("details", ""), font=FONT_BODY,
              fill=row_fill, border=BORDER, alignment=WRAP)

    _set_col_widths(ws, [10, 40, 70])


# ── main ──────────────────────────────────────────────────────────────────────

def generate(findings_path, output_path=None):
    with open(findings_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    if not output_path:
        hostname = data.get("metadata", {}).get("hostname", "dfir")
        report_dir = os.path.join(os.path.dirname(findings_path),
                                  f"{hostname}_DFIR_Report")
        os.makedirs(report_dir, exist_ok=True)
        output_path = os.path.join(report_dir, f"{hostname}_DFIR_Report.xlsx")

    wb = openpyxl.Workbook()

    _sheet_summary(wb, data)
    _sheet_network(wb, data)
    _sheet_processes(wb, data)
    _sheet_persistence(wb, data)
    _sheet_software(wb, data)
    _sheet_logins(wb, data)
    _sheet_file_activity(wb, data)
    _sheet_shell_history(wb, data)
    _sheet_browser_history(wb, data)
    _sheet_event_logs(wb, data)   # only added if event_logs list is non-empty
    _sheet_iocs(wb, data)
    _sheet_recommendations(wb, data)

    wb.save(output_path)
    print(f"Report saved: {output_path}")
    return output_path


def main():
    parser = argparse.ArgumentParser(
        description="Generate formatted DFIR report xlsx from JSON findings")
    parser.add_argument("findings", help="Path to Claude's JSON findings file")
    parser.add_argument("--out", help="Output .xlsx path (default: alongside findings file)")
    args = parser.parse_args()

    if not os.path.exists(args.findings):
        sys.exit(f"File not found: {args.findings}")

    generate(args.findings, args.out)


if __name__ == "__main__":
    main()
